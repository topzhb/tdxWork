#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
概念板块HTML报告生成器 - V2.0
创建日期：2026-03-23
更新日期：2026-04-02

V2.0 升级内容：
1. 趋势Tab重构：日线/三日/周线三个独立Tab，各自使用对应周期的妖线满足数
2. 每日backfill同时写入三个周期数据
3. run_id格式改为 {YYYYMMDD_HIST}_{PERIOD}，period='daily'|'3day'|'weekly'
4. 数据库 period 字段标识各周期数据
5. --date 参数（推荐）：指定报告日期，自动查找该日期的三个周期 run_id
6. --run-id 保留向后兼容

V1.5 升级内容：
1. 板块涨幅数据来源：从板块指数获取（而非个股平均）
2. 支持盘中实时数据：通过--intraday参数使用通达信缓存数据
3. 通达信数据路径可配置

特点：
- 完全从 SQLite 数据库读取（t_sector_stat / t_stock_calc / t_daily_report）
- 新增"趋势跟踪"Tab：展示历史各日期满足率变化
- 满足条件个股明细从去重后的 t_stock_calc 直接读取
- 市场总满足数趋势图：支持1个月/3个月/1年/全部时间范围切换
- 板块双击事件：双击表格中的板块名称弹出该板块历史趋势图
- 表格日期筛选：支持二周/一个月/三个月/全部数据切换
- 输出：ConceptReport/concept_report_YYYYMMDD.html
"""

import sqlite3
import os
import sys
import json
import struct
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 使用相对路径 - 适应当前工作目录
DB_FILE     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "db", "concept_weekly.db")
REPORT_DIR  = "ConceptReport"

# 通达信数据路径（从 tool_config 集中配置）
from tool_config import VIPDOC_DIR as TDX_VIPDOC_DIR
TDX_CACHE_DIR = os.path.join(os.path.dirname(TDX_VIPDOC_DIR), "T0002", "cache")  # 盘中缓存数据目录

# 获取输出文件名（在main函数中动态设置）
OUTPUT_FILE = None  # 将在main函数中设置

# 全局配置
USE_INTRADAY = False  # 是否使用盘中实时数据


def check_database_file():
    """检查数据库文件是否存在"""
    if not os.path.exists(DB_FILE):
        print(f"[ERROR] 数据库文件不存在: {DB_FILE}")
        print("当前工作目录:", os.getcwd())
        print("请确保在项目根目录下运行此脚本")
        return False
    return True


def check_report_directory():
    """检查报告目录是否存在，不存在则创建"""
    try:
        os.makedirs(REPORT_DIR, exist_ok=True)
        return True
    except Exception as e:
        print(f"[ERROR] 无法创建报告目录 {REPORT_DIR}: {e}")
        return False


def get_run_id_for_date(conn, date_str, period='weekly'):
    """
    Get run_id from date and period.
    date_str: 'YYYY-MM-DD' or 'YYYYMMDD'
    period: 'daily' | '3day' | 'weekly'
    """
    date_code = date_str.replace('-', '')
    run_id = date_code + "_HIST_" + period.upper()
    exists = conn.execute(
        "SELECT 1 FROM t_run_log WHERE run_id = ?", (run_id,)
    ).fetchone()
    return run_id if exists else None


def get_latest_run_id(conn):
    """Get latest weekly run_id"""
    row = conn.execute(
        "SELECT run_id FROM t_run_log WHERE run_id LIKE '%_WEEKLY' ORDER BY run_time DESC LIMIT 1"
    ).fetchone()
    return row[0] if row else None


def load_summary(conn, run_id):
    row = conn.execute("""
        SELECT version, total_sectors, total_stocks, unique_stocks,
               analyzed_count, satisfied_count, satisfied_rate, run_time
        FROM t_run_log WHERE run_id=?
    """, (run_id,)).fetchone()
    if not row:
        return {}
    return {
        'version':         row[0],
        'sector_count':    row[1],
        'total_stocks':    row[2],
        'unique_stocks':   row[3],
        'analyzed_count':  row[4],
        'satisfied_count': row[5],
        'satisfied_rate':  row[6],
        'run_time':        row[7],
    }


def load_historical_data(conn, days=90):
    """
    获取最近days天的历史数据，返回三个周期的满足数趋势数据
    从t_run_log表中获取数据，每个日期每个周期只取最新的一条记录
    返回格式: {'daily': [{'date': '2026-01-01', 'value': 54}, ...], ...}
    """
    from datetime import datetime, timedelta
    
    # 计算起始日期
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    start_date_str = start_date.strftime('%Y%m%d')  # 格式：20260101
    
    # 查询三个周期的历史数据，每个日期每个周期只取最新的一条
    query = """
        SELECT 
            period,
            report_date,
            satisfied_count
        FROM (
            SELECT 
                CASE 
                    WHEN run_id LIKE '%_DAILY' THEN 'daily'
                    WHEN run_id LIKE '%_3DAY' THEN '3day'
                    WHEN run_id LIKE '%_WEEKLY' THEN 'weekly'
                    ELSE 'unknown'
                END as period,
                -- 从run_id中提取日期：20260403_HIST_DAILY → 2026-04-03
                CASE 
                    WHEN SUBSTR(run_id, 1, 8) LIKE '2026____' THEN 
                        SUBSTR(run_id, 1, 4) || '-' || SUBSTR(run_id, 5, 2) || '-' || SUBSTR(run_id, 7, 2)
                    ELSE NULL
                END as report_date,
                satisfied_count,
                ROW_NUMBER() OVER (PARTITION BY 
                    CASE 
                        WHEN run_id LIKE '%_DAILY' THEN 'daily'
                        WHEN run_id LIKE '%_3DAY' THEN '3day'
                        WHEN run_id LIKE '%_WEEKLY' THEN 'weekly'
                    END, 
                    -- 按从run_id提取的日期分区
                    CASE 
                        WHEN SUBSTR(run_id, 1, 8) LIKE '2026____' THEN 
                            SUBSTR(run_id, 1, 4) || '-' || SUBSTR(run_id, 5, 2) || '-' || SUBSTR(run_id, 7, 2)
                        ELSE NULL
                    END
                    ORDER BY run_time DESC
                ) as rn
            FROM t_run_log 
            WHERE -- 按run_id中的日期过滤
                (run_id LIKE '2026%_HIST_DAILY' OR run_id LIKE '2026%_HIST_3DAY' OR run_id LIKE '2026%_HIST_WEEKLY')
                AND run_id >= ? || '_HIST_DAILY'  -- 比较日期部分
        ) t
        WHERE t.rn = 1 AND t.report_date IS NOT NULL
        ORDER BY report_date
    """
    
    cursor = conn.execute(query, (start_date_str,))
    rows = cursor.fetchall()
    
    # 组织数据
    data = {'daily': [], '3day': [], 'weekly': []}
    
    for period, report_date, satisfied_count in rows:
        if period not in data:
            continue
            
        # 处理数值（根据需求）
        value = satisfied_count
        if period == 'weekly':
            # 周线数据除以4
            value = value / 4
        
        # 上限250
        value = min(value, 250)
        
        data[period].append({
            'date': report_date,
            'value': round(value, 2)
        })
    
    return data


def load_sector_stats(conn, run_id):
    rows = conn.execute("""
        SELECT s.sector_name, ss.total_stocks,
               ss.analyzed_count, ss.satisfied_count, ss.satisfied_rate,
               ss.sector_code
        FROM t_sector_stat ss
        JOIN t_sector s ON ss.sector_code = s.sector_code
        WHERE ss.run_id = ?
        ORDER BY ss.satisfied_count DESC
    """, (run_id,)).fetchall()
    # 按新排序重新分配排名
    return [{
        'rank': idx + 1,
        'name': r[0], 'total': r[1],
        'analyzed': r[2], 'satisfied': r[3],
        'rate': r[4], 'sector_code': r[5],
    } for idx, r in enumerate(rows)]


def load_satisfied_stocks(conn, run_id, sector_code, top_n=20):
    """通过 t_sector_stock 关联 t_stock_calc 获取满足条件个股"""
    rows = conn.execute("""
        SELECT sc.stock_code,
               COALESCE(st.stock_name, '') as stock_name,
               sc.angle, sc.zhusheng1, sc.close_price, sc.x_val
        FROM t_stock_calc sc
        JOIN t_sector_stock ss ON sc.stock_code = ss.stock_code
                               AND ss.sector_code = ?
        LEFT JOIN t_stock st ON sc.stock_code = st.stock_code
        WHERE sc.run_id = ? AND sc.is_satisfied = 1
        ORDER BY sc.angle DESC
        LIMIT ?
    """, (sector_code, run_id, top_n)).fetchall()
    return [{'code': r[0], 'name': r[1], 'angle': r[2],
             'zhusheng1': r[3], 'close': r[4], 'x': r[5]} for r in rows]


def load_trend_data(conn, period='weekly', top_n_sectors=20):
    """
    从 t_daily_report 读取历史各日期的板块满足率（趋势跟踪，仅包含交易日）
    返回: dates列表, sectors列表, data矩阵
    """
    # 获取所有已记录日期（升序，按指定period过滤）
    all_dates = [r[0] for r in conn.execute(
        "SELECT DISTINCT report_date FROM t_daily_report WHERE period=? ORDER BY report_date ASC",
        (period,)
    ).fetchall()]

    # 过滤掉周末日期
    dates = []
    for date_str in all_dates:
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            # 跳过周末（周六=5, 周日=6）
            if date_obj.weekday() < 5:
                dates.append(date_str)
        except:
            # 如果日期格式异常，保留该日期
            dates.append(date_str)

    if not dates:
        return [], {}, {}

    # 取最新一天排名 TOP N 的板块名（按period过滤）
    latest_date = dates[-1]
    top_sectors = conn.execute("""
        SELECT sector_code, sector_name, rank_no
        FROM t_daily_report
        WHERE report_date = ? AND period = ?
        ORDER BY rank_no ASC LIMIT ?
    """, (latest_date, period, top_n_sectors)).fetchall()

    sector_codes = [r[0] for r in top_sectors]
    sector_names = {r[0]: r[1] for r in top_sectors}

    # 构建 {sector_code: {date: rate}} 的矩阵（仅包含过滤后的日期）
    data = {code: {} for code in sector_codes}
    if sector_codes and dates:
        rows = conn.execute("""
            SELECT sector_code, report_date, satisfied_rate
            FROM t_daily_report
            WHERE sector_code IN ({}) AND report_date IN ({}) AND period = ?
            ORDER BY report_date ASC
        """.format(','.join(['?'] * len(sector_codes)), ','.join(['?'] * len(dates))),
                    sector_codes + dates + [period]).fetchall()

        for r in rows:
            if r[0] in data:
                data[r[0]][r[1]] = r[2]

    return dates, sector_names, data


def load_global_trend(conn, period='weekly'):
    """
    从 t_run_log 读取总体满足数日线趋势（仅包含交易日，过滤周末）
    period: 过滤指定周期的 run_id
    返回: dates列表, satisfied_counts列表, satisfaction_rates列表
    """
    period_pattern = f"%_{period.upper()}"
    rows = conn.execute("""
        SELECT run_id, satisfied_count, satisfied_rate, note
        FROM t_run_log
        WHERE note LIKE '历史回溯%' AND run_id LIKE ?
        ORDER BY run_id ASC
    """, (period_pattern,)).fetchall()

    if not rows:
        return [], [], []

    # 从 note 中提取日期 (格式: "历史回溯：2026-03-01 [三日]" 或旧格式 "历史回溯：2026-03-01")
    dates = []
    sat_counts = []
    sat_rates  = []
    for run_id, sat_cnt, sat_rate, note in rows:
        # 提取日期部分：去掉中括号及之后的内容
        raw = note.split('：')[1] if '：' in note else run_id[:8]
        date_str = raw.split('[')[0].strip()  # "2026-04-02 [三日]" → "2026-04-02"

        # 将日期转换为 datetime 以检查是否为周末
        if len(date_str) == 8:
            # YYYYMMDD格式
            date_obj = datetime.strptime(date_str, '%Y%m%d')
        elif '-' in date_str:
            # YYYY-MM-DD格式
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        else:
            date_obj = None

        # 跳过周末数据（周六=5, 周日=6）
        if date_obj and date_obj.weekday() >= 5:
            continue

        # 将 YYYYMMDD 转换为 YYYY-MM-DD
        if len(date_str) == 8:
            formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
            dates.append(formatted_date)
        else:
            dates.append(date_str)
        sat_counts.append(sat_cnt)
        sat_rates.append(sat_rate)

    return dates, sat_counts, sat_rates


def load_sector_multi_period_trend(conn, sector_code):
    """获取指定板块的三个周期历史满足数数据（用于弹窗中的三周期趋势图）
    
    从 t_daily_report 表查询该板块在 daily/3day/weekly 三个周期下的
    历史满足数数据，返回按日期合并的数据结构。
    
    返回格式: {
        'dates': ['2026-03-01', '2026-03-02', ...],
        'daily': [10, 12, ...],
        '3day': [5, 6, ...],
        'weekly': [40, 45, ...]  # 已除以4
    }
    """
    rows = conn.execute("""
        SELECT report_date, period, satisfied_count
        FROM t_daily_report
        WHERE sector_code = ?
          AND period IN ('daily', '3day', 'weekly')
        ORDER BY report_date ASC, period ASC
    """, (sector_code,)).fetchall()
    
    # 按日期收集所有周期的数据
    date_map = {}  # {date_str: {'daily': val, '3day': val, 'weekly': val}}
    dates_ordered = []
    
    for report_date, period, sat_count in rows:
        if report_date not in date_map:
            date_map[report_date] = {}
            dates_ordered.append(report_date)
        
        # 处理数值：周线数据除以4（与汇总趋势图一致）
        value = sat_count
        if period == 'weekly':
            value = sat_count / 4
        
        date_map[report_date][period] = round(value, 2)
    
    # 构建返回结果
    result = {
        'dates': dates_ordered,
        'daily': [],
        '3day': [],
        'weekly': []
    }
    
    for d in dates_ordered:
        dm = date_map[d]
        result['daily'].append(dm.get('daily', None))
        result['3day'].append(dm.get('3day', None))
        result['weekly'].append(dm.get('weekly', None))
    
    return result


def load_sector_trend(conn, sector_code, period='weekly'):
    """获取指定板块的历史趋势数据（返回板块满足数和总比满足率）"""
    dates = []
    sat_counts = []
    sat_rates = []

    # 获取该板块的历史数据以及对应的每日总满足数（按period过滤）
    rows = conn.execute("""
        SELECT
            ds.report_date,
            ds.satisfied_count,
            (SELECT satisfied_count FROM t_run_log
             WHERE run_id = (
               SELECT run_id FROM t_daily_report
               WHERE report_date = ds.report_date AND period = ds.period
               LIMIT 1
             )
            ) as total_satisfied
        FROM t_daily_report ds
        WHERE ds.sector_code = ? AND ds.period = ?
        ORDER BY ds.report_date ASC
    """, (sector_code, period)).fetchall()

    for date_str, sat_cnt, total_sat in rows:
        # 跳过周末数据
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            if date_obj.weekday() >= 5:  # 周末跳过
                continue
        except:
            pass

        dates.append(date_str)
        sat_counts.append(sat_cnt)

        # 计算总比满足率
        if total_sat and total_sat > 0:
            total_rate = (sat_cnt / total_sat) * 100
        else:
            total_rate = 0
        sat_rates.append(total_rate)

    return dates, sat_counts, sat_rates


def load_daily_top20(conn, period, dates):
    """
    为每一天获取TOP20板块（基于满足数排序，按period过滤）
    返回: {日期: {total_satisfied: 总数, sectors: [{rank, name, total_rate, code, satisfied_count}, ...]}}
    """
    daily_top20 = {}

    for date_str in dates:
        # 获取该日期的板块数据（从t_daily_report，按period过滤）
        rows = conn.execute("""
            SELECT
                ds.sector_code,
                s.sector_name,
                ds.satisfied_count,
                (SELECT satisfied_count FROM t_run_log
                 WHERE run_id = (
                   SELECT run_id FROM t_daily_report
                   WHERE report_date = ? AND period = ?
                   LIMIT 1
                 )
                ) as total_satisfied
            FROM t_daily_report ds
            JOIN t_sector s ON ds.sector_code = s.sector_code
            WHERE ds.report_date = ? AND ds.period = ?
            ORDER BY ds.satisfied_count DESC
            LIMIT 20
        """, (date_str, period, date_str, period)).fetchall()

        # 计算总比满足率并排序
        sectors = []
        total_satisfied = 0
        if rows:
            total_satisfied = rows[0][3]  # 该日期的总满足数
            if total_satisfied == 0:
                total_satisfied = 1  # 防零除

            for code, name, satisfied, total in rows:
                total_rate = (satisfied / total_satisfied) * 100 if total > 0 else 0
                sectors.append({
                    'code': code,
                    'name': name,
                    'total_rate': total_rate,
                    'satisfied_count': satisfied
                })

            # 按总比满足率排序
            sectors.sort(key=lambda x: x['total_rate'], reverse=True)

            # 重新分配排名
            for idx, sector in enumerate(sectors):
                sector['rank'] = idx + 1

        daily_top20[date_str] = {
            'total_satisfied': total_satisfied,
            'sectors': sectors
        }

    return daily_top20


def is_limit_up(stock_code, daily_change):
    """
    判断是否为涨停
    :param stock_code: 股票代码
    :param daily_change: 当日涨幅百分比
    :return: True/False
    """
    # 创业板、科创板 - 20%涨停
    if stock_code.startswith(('30', '688')):
        return daily_change >= 19.9
    # 主板 - 10%涨停
    else:
        return daily_change >= 9.9


def read_sector_index_change(sector_code, use_intraday=False):
    """
    从通达信板块指数读取当日涨幅
    
    Args:
        sector_code: 板块代码 (如 '880507')
        use_intraday: 是否使用盘中缓存数据
    
    Returns:
        涨幅百分比 (如 2.5 表示上涨2.5%)，如果读取失败返回 None
    """
    # 板块指数文件路径
    # 板块指数都在上海市场
    index_file = f"sh{sector_code}"
    
    if use_intraday and os.path.exists(TDX_CACHE_DIR):
        # 使用盘中缓存数据
        cache_file = os.path.join(TDX_CACHE_DIR, f"{index_file}.~~~day")
        if os.path.exists(cache_file):
            return parse_tdx_cache_file(cache_file)
    
    # 使用标准日线数据
    day_file = os.path.join(TDX_VIPDOC_DIR, "sh", "lday", f"{index_file}.day")
    if os.path.exists(day_file):
        return parse_tdx_day_file(day_file)
    
    return None


def parse_tdx_cache_file(filepath):
    """
    解析通达信缓存文件 (.~~~day)
    格式：12字节头部 + N条记录（日期4字节 + 开高低收4字节浮点 + 成交额4字节 + 成交量4字节）
    
    修正：涨幅计算使用 (今日收盘 - 昨日收盘) / 昨日收盘
    """
    try:
        with open(filepath, 'rb') as f:
            data = f.read()
        
        if len(data) < 32:
            return None
        
        # 收集所有有效记录
        records = []
        for i in range(12, len(data) - 20, 4):
            try:
                date_val = struct.unpack('<I', data[i:i+4])[0]
                year = date_val // 10000
                if 2020 <= year <= 2030:
                    if i + 20 <= len(data):
                        values = struct.unpack('<Iffff', data[i:i+20])
                        records.append(values)
            except:
                continue
        
        if len(records) < 2:
            # 不足两条记录，使用最后一条的开盘收盘计算（fallback）
            if records:
                date_val, open_p, high, low, close = records[-1]
                if open_p > 0:
                    return (close - open_p) / open_p * 100
            return None
        
        # 使用最后两条记录计算涨幅（今日 vs 昨日）
        prev_date, prev_open, prev_high, prev_low, prev_close = records[-2]
        last_date, last_open, last_high, last_low, last_close = records[-1]
        
        # 计算涨幅（相对于昨日收盘）
        if prev_close > 0:
            change_pct = (last_close - prev_close) / prev_close * 100
            return change_pct
        
        return None
    except Exception as e:
        print(f"解析缓存文件出错 {filepath}: {e}")
        return None


def parse_tdx_day_file(filepath):
    """
    解析通达信标准日线文件 (.day)
    格式：每条记录32字节
    - date: 4字节 (YYYYMMDD)
    - open: 4字节 (价格*100)
    - high: 4字节 (价格*100)
    - low: 4字节 (价格*100)
    - close: 4字节 (价格*100)
    - amount: 4字节 (成交金额/1000)
    - volume: 4字节 (成交量/100)
    - reserved: 4字节 (保留)
    
    修正：涨幅计算使用 (今日收盘 - 昨日收盘) / 昨日收盘
    """
    try:
        with open(filepath, 'rb') as f:
            data = f.read()
        
        if len(data) < 64:  # 至少需要两条记录
            return None
        
        record_size = 32
        factor = 100.0
        
        # 读取最后一条记录（今日数据）
        last_record_start = len(data) - record_size
        last_record = data[last_record_start:last_record_start+record_size]
        last_values = struct.unpack('<IIIIIIII', last_record)
        last_date, last_open, last_high, last_low, last_close, last_amount, last_volume, last_reserved = last_values
        
        # 读取倒数第二条记录（昨日数据）
        prev_record_start = len(data) - record_size * 2
        prev_record = data[prev_record_start:prev_record_start+record_size]
        prev_values = struct.unpack('<IIIIIIII', prev_record)
        prev_date, prev_open, prev_high, prev_low, prev_close, prev_amount, prev_volume, prev_reserved = prev_values
        
        # 转换价格
        last_close_price = last_close / factor
        prev_close_price = prev_close / factor
        
        # 计算涨幅（相对于昨日收盘）
        if prev_close_price > 0:
            change_pct = (last_close_price - prev_close_price) / prev_close_price * 100
            return change_pct
        
        return None
    except Exception as e:
        print(f"解析日线文件出错 {filepath}: {e}")
        return None


def load_sector_change_stats(conn, latest_run_id, prev_run_id=None, use_intraday=False):
    """
    加载板块涨跌统计数据（用于左栏表格）
    返回: [{sector_name, limit_up_count, analyzed_count, limit_up_ratio, 
            avg_satisfied_change, avg_all_change, change_ratio, change_arrow}, ...]
    默认按板块涨幅(avg_all_change)降序排列
    
    V1.5更新：
    - avg_all_change 从板块指数获取（而非个股平均）
    - 支持 use_intraday 参数获取盘中实时数据
    """
    # 获取最新日期的数据
    latest_date = conn.execute(
        "SELECT calc_date FROM t_stock_calc WHERE run_id = ? LIMIT 1",
        (latest_run_id,)
    ).fetchone()
    if not latest_date:
        return []
    latest_date = latest_date[0]

    # 获取前一日数据（用于对比）
    prev_date = None
    if prev_run_id:
        prev_row = conn.execute(
            "SELECT calc_date FROM t_stock_calc WHERE run_id = ? LIMIT 1",
            (prev_run_id,)
        ).fetchone()
        if prev_row:
            prev_date = prev_row[0]

    # 查询所有板块
    sectors = conn.execute("SELECT sector_code, sector_name FROM t_sector").fetchall()

    # 预加载虚拟板块配置，计算所有子板块的指数涨幅缓存
    virtual_config = {}
    sub_sector_index_changes = {}  # {子板块代码: 板块指数涨幅}
    try:
        vs_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'virtual_sectors.json')
        if os.path.exists(vs_file):
            with open(vs_file, 'r', encoding='utf-8') as f:
                vs_data = json.load(f)
            for vs in vs_data.get('virtual_sectors', []):
                virtual_config[vs['code']] = [sub['code'] for sub in vs.get('sub_sectors', [])]
                for sub_code in vs.get('sub_sectors', []):
                    if sub_code['code'] not in sub_sector_index_changes:
                        chg = read_sector_index_change(sub_code['code'], use_intraday=use_intraday)
                        sub_sector_index_changes[sub_code['code']] = chg
    except Exception as e:
        print(f"[WARN] 加载虚拟板块配置失败: {e}")

    result = []
    for sector_code, sector_name in sectors:
        # 获取该板块最新日期的所有股票数据
        rows = conn.execute("""
            SELECT sc.stock_code, sc.daily_change, sc.is_satisfied
            FROM t_stock_calc sc
            JOIN t_sector_stock ss ON sc.stock_code = ss.stock_code
            WHERE ss.sector_code = ? AND sc.run_id = ?
        """, (sector_code, latest_run_id)).fetchall()

        if not rows:
            continue

        # 统计涨停数和满足条件股票的涨幅
        limit_up_count = 0
        satisfied_changes = []

        for stock_code, daily_change, is_satisfied in rows:
            if daily_change is None:
                daily_change = 0
            if is_limit_up(stock_code, daily_change):
                limit_up_count += 1
            if is_satisfied:
                satisfied_changes.append(daily_change)

        analyzed_count = len(rows)
        limit_up_ratio = (limit_up_count / analyzed_count * 100) if analyzed_count > 0 else 0

        # 判断是否为虚拟板块（代码以888开头）
        is_virtual = sector_code.startswith('888')

        # V1.5：从板块指数获取板块涨幅
        if is_virtual:
            # 虚拟板块涨幅 = 关联子板块指数涨幅的平均值
            sub_codes = virtual_config.get(sector_code, [])
            valid_changes = [sub_sector_index_changes[c] for c in sub_codes
                             if c in sub_sector_index_changes and sub_sector_index_changes[c] is not None]
            avg_all_change = sum(valid_changes) / len(valid_changes) if valid_changes else None
        else:
            sector_index_change = read_sector_index_change(sector_code, use_intraday=use_intraday)

            # 如果无法从板块指数获取，则使用个股平均涨幅作为后备
            if sector_index_change is not None:
                avg_all_change = sector_index_change
            else:
                # 后备方案：计算个股平均涨幅
                total_change = sum(r[1] for r in rows if r[1] is not None)
                avg_all_change = total_change / analyzed_count if analyzed_count > 0 else 0
        
        avg_satisfied_change = sum(satisfied_changes) / len(satisfied_changes) if satisfied_changes else 0
        change_ratio = (avg_satisfied_change / avg_all_change) if (avg_all_change is not None and avg_all_change != 0) else 0

        # 计算与前一日的对比箭头
        change_arrow = ''
        if prev_run_id:
            # 获取前一日的板块指数涨幅（虚拟板块从prev_run_id的个股数据中取子板块平均）
            if is_virtual:
                # 虚拟板块前日涨幅：取子板块在prev_run_id下的个股平均涨幅均值
                sub_codes = virtual_config.get(sector_code, [])
                prev_sub_changes = []
                for sub_code in sub_codes:
                    sub_rows = conn.execute("""
                        SELECT sc.daily_change FROM t_stock_calc sc
                        JOIN t_sector_stock ss ON sc.stock_code = ss.stock_code
                        WHERE ss.sector_code = ? AND sc.run_id = ?
                    """, (sub_code, prev_run_id)).fetchall()
                    if sub_rows:
                        valid = [r[0] for r in sub_rows if r[0] is not None]
                        if valid:
                            prev_sub_changes.append(sum(valid) / len(valid))
                prev_sector_index_change = sum(prev_sub_changes) / len(prev_sub_changes) if prev_sub_changes else None
            else:
                prev_sector_index_change = read_sector_index_change(sector_code, use_intraday=False)

            if prev_sector_index_change is not None:
                # 使用板块指数涨幅对比
                if avg_all_change > prev_sector_index_change:
                    change_arrow = '↑'
                elif avg_all_change < prev_sector_index_change:
                    change_arrow = '↓'
            else:
                # 后备方案：使用涨幅比对比
                prev_rows = conn.execute("""
                    SELECT sc.stock_code, sc.daily_change, sc.is_satisfied
                    FROM t_stock_calc sc
                    JOIN t_sector_stock ss ON sc.stock_code = ss.stock_code
                    WHERE ss.sector_code = ? AND sc.run_id = ?
                """, (sector_code, prev_run_id)).fetchall()

                if prev_rows:
                    prev_total_change = 0
                    prev_satisfied_changes = []
                    for stock_code, daily_change, is_satisfied in prev_rows:
                        if daily_change is None:
                            daily_change = 0
                        prev_total_change += daily_change
                        if is_satisfied:
                            prev_satisfied_changes.append(daily_change)

                    prev_analyzed_count = len(prev_rows)
                    prev_avg_all_change = prev_total_change / prev_analyzed_count if prev_analyzed_count > 0 else 0
                    prev_avg_satisfied_change = sum(prev_satisfied_changes) / len(prev_satisfied_changes) if prev_satisfied_changes else 0
                    prev_change_ratio = (prev_avg_satisfied_change / prev_avg_all_change) if prev_avg_all_change != 0 else 0

                    if change_ratio > prev_change_ratio:
                        change_arrow = '↑'
                    elif change_ratio < prev_change_ratio:
                        change_arrow = '↓'

        result.append({
            'sector_name': sector_name,
            'sector_code': sector_code,
            'is_virtual': is_virtual,
            'limit_up_count': limit_up_count,
            'analyzed_count': analyzed_count,
            'limit_up_ratio': limit_up_ratio,
            'avg_satisfied_change': avg_satisfied_change,
            'avg_all_change': avg_all_change if avg_all_change is not None else 0,
            'change_ratio': change_ratio,
            'change_arrow': change_arrow
        })

    # 默认按板块涨幅(avg_all_change)降序排列，虚拟板块排在最后
    result.sort(key=lambda x: (1 if x.get('is_virtual') else 0, -x['avg_all_change']))

    return result  # 返回所有板块，由调用方决定显示数量


def build_trend_html(conn, period='weekly', full_mode=False, use_intraday=False):
    """生成趋势跟踪区块 HTML（Chart.js 混合图表 + 每日TOP20表格 + 日期筛选）
    period: 'daily' | '3day' | 'weekly'
    full_mode: True=加载全部数据, False=默认只加载最近三个月(60个交易日)
    use_intraday: True=使用盘中实时数据, False=使用收盘数据
    """
    all_dates_raw, _, _ = load_trend_data(conn, period=period, top_n_sectors=20)

    if not all_dates_raw or len(all_dates_raw) < 1:
        return '<div class="no-trend">暂无历史数据，运行多次分析后将显示趋势图</div>'

    if len(all_dates_raw) == 1:
        return f'<div class="no-trend">当前仅有 {all_dates_raw[0]} 一天的数据，继续每日运行后将显示趋势变化</div>'

    # 限制加载范围：full_mode=False 时只取最近三个月（60个交易日）
    if full_mode:
        all_dates = all_dates_raw
    else:
        all_dates = all_dates_raw[-60:] if len(all_dates_raw) >= 60 else all_dates_raw

    # 为每一天获取TOP20板块
    daily_top20 = load_daily_top20(conn, period, all_dates)

    # 获取最新和前一个run_id，用于加载板块涨跌统计（按period过滤）
    period_pattern = f"%_{period.upper()}"
    run_ids = conn.execute(
        "SELECT run_id FROM t_run_log WHERE run_id LIKE ? ORDER BY run_time DESC LIMIT 2",
        (period_pattern,)
    ).fetchall()
    latest_run_id = run_ids[0][0] if len(run_ids) > 0 else None
    prev_run_id = run_ids[1][0] if len(run_ids) > 1 else None

    # 加载板块涨跌统计数据（传入use_intraday参数）
    sector_change_stats = []
    if latest_run_id:
        sector_change_stats = load_sector_change_stats(conn, latest_run_id, prev_run_id, use_intraday=use_intraday)

    # 获取最新日期的TOP20板块代码（用于左侧表格底色判断）
    latest_date = all_dates[-1] if all_dates else None
    top20_sector_codes = set()
    if latest_date:
        latest_daily_data = daily_top20.get(latest_date, {})
        for sector in latest_daily_data.get('sectors', []):
            top20_sector_codes.add(sector['code'])

    # 生成表格HTML生成函数（用于JavaScript动态调用）
    def generate_table_html(dates):
        """生成指定日期范围的表格HTML"""
        # 生成表头（排名列 + 日期列）
        date_headers = ''.join(f'<th style="min-width:120px;">{d[5:]}</th>' for d in dates)  # 只显示 MM-DD
        thead = f'<tr><th style="width:60px;">排名</th>{date_headers}</tr>'

        # 生成表格内容 - 按排名（1-20）为行，每天为列
        tbody = ''

        # 第一行：总数（满足条件的总数）
        total_row = '<tr style="background:#e8f4fd; font-weight:bold;">'
        total_row += '<td class="rank-cell">总数</td>'
        for date_idx, date_str in enumerate(dates):
            daily_data = daily_top20.get(date_str, {})
            total_count = daily_data.get('total_satisfied', 0)

            # 计算与后一天（右侧列，即实际上的前一天）的变化
            total_arrow = ''
            if date_idx < len(dates) - 1:
                next_date = dates[date_idx + 1]
                next_data = daily_top20.get(next_date, {})
                next_total = next_data.get('total_satisfied', 0)

                if total_count > next_total:
                    total_arrow = ' <span style="color:#e84040">▲</span>'
                elif total_count < next_total:
                    total_arrow = ' <span style="color:#1cb85e">▼</span>'

            total_row += f'<td style="text-align:center;">{total_count}{total_arrow}</td>'
        total_row += '</tr>'

        # 建立映射：{排名: {日期: 板块数据}}
        rank_data = {}
        for date_idx, date_str in enumerate(dates):
            daily_data = daily_top20.get(date_str, {})
            sectors = daily_data.get('sectors', [])
            for sector in sectors:
                rank = sector['rank']
                if rank not in rank_data:
                    rank_data[rank] = {}
                rank_data[rank][date_str] = {
                    'name': sector['name'],
                    'rate': sector['total_rate'],
                    'code': sector['code'],
                    'satisfied_count': sector['satisfied_count']
                }

        # 按排名1-20生成行
        for rank in range(1, 21):
            if rank not in rank_data:
                continue

            rank_row = f'<tr><td class="rank-cell">{rank}</td>'

            for date_idx, date_str in enumerate(dates):
                sector = rank_data[rank].get(date_str)

                if not sector:
                    # 该排名当天没有数据（可能少于20个板块）
                    rank_row += '<td>-</td>'
                    continue

                name = sector['name']
                rate = sector['rate']
                code = sector['code']
                satisfied_count = sector.get('satisfied_count', 0)

                # 计算个股个数与后一天（右侧列，即实际上的前一天）的变化
                count_arrow = ''
                if date_idx < len(dates) - 1:
                    next_date = dates[date_idx + 1]
                    next_data = daily_top20.get(next_date, {})
                    next_sectors = next_data.get('sectors', [])

                    # 在后一天的所有板块中查找当前板块
                    next_sector_found = None
                    for ns in next_sectors:
                        if ns['code'] == code:
                            next_sector_found = ns
                            break

                    if next_sector_found:
                        next_count = next_sector_found.get('satisfied_count', 0)
                        if satisfied_count > next_count:
                            count_arrow = '<span style="color:#e84040; font-size:0.85em;">↑</span>'  # 上升-红色
                        elif satisfied_count < next_count:
                            count_arrow = '<span style="color:#1cb85e; font-size:0.85em;">↓</span>'  # 下降-绿色

                # 板块名称后面增加"(满足数+升降箭头)"，并添加点击和双击事件
                name_with_count = f'<span class="sector-name" data-sector-code="{code}" onclick="highlightSector(\'{code}\')" ondblclick="showSectorTrend(\'{code}\', \'{name}\')">{name} <span style="color:#888; font-size:0.85em;">({satisfied_count}{count_arrow})</span></span>'

                # 计算与后一天（右侧列，即实际上的前一天）的变化（基于板块代码，而非排名位置）
                arrow = ''
                cls = ''
                if date_idx < len(dates) - 1:
                    next_date = dates[date_idx + 1]
                    next_data = daily_top20.get(next_date, {})
                    next_sectors = next_data.get('sectors', [])

                    # 在后一天的所有板块中查找当前板块
                    next_sector_found = None
                    for ns in next_sectors:
                        if ns['code'] == code:
                            next_sector_found = ns
                            break

                    if next_sector_found:
                        # 情况1：后一天在TOP20中，比较数值变化
                        next_rate = next_sector_found['total_rate']
                        diff = rate - next_rate
                        if diff > 0.5:
                            arrow = '<span class="trend-arrow"> ▲</span>'
                            cls = 'trend-up'
                        elif diff < -0.5:
                            arrow = '<span class="trend-arrow"> ▼</span>'
                            cls = 'trend-down'
                    else:
                        # 情况2：后一天不在TOP20中，新进入表格，采用上升标识
                        arrow = '<span class="trend-arrow"> ▲</span>'
                        cls = 'trend-up'

                # 生成单元格内容（板块名称 + 满足数+升降箭头 + 总比满足率，换行显示）
                cell_content = f'<div class="sector-name" data-sector-code="{code}" onclick="highlightSector(\'{code}\')" ondblclick="showSectorTrend(\'{code}\', \'{name}\')">{name} <span style="color:#888; font-size:0.85em;">({satisfied_count}){count_arrow}</span></div><div class="rate-value">{rate:.2f}%{arrow}</div>'
                rank_row += f'<td class="{cls}" data-sector-code="{code}">{cell_content}</td>'

            rank_row += '</tr>'
            tbody += rank_row

        # 完整表格HTML = 表头 + 总数行 + 排名行
        return thead + total_row + tbody

    # 生成表格日期范围切片（倒序排列，最新日期在前）
    two_week_dates = all_dates[-10:] if len(all_dates) >= 10 else all_dates  # 最近二周（10个交易日）
    one_month_dates = all_dates[-20:] if len(all_dates) >= 20 else all_dates  # 最近一个月（20个交易日）
    one_week_dates = all_dates[-5:] if len(all_dates) >= 5 else all_dates  # 最近一周（5个交易日）
    three_month_dates = all_dates  # 三个月 = 当前已加载的全部数据

    # 将表格HTML作为JavaScript变量存储（日期倒序排列）
    table_html_two_week = generate_table_html(two_week_dates[::-1])
    table_html_one_month = generate_table_html(one_month_dates[::-1])
    table_html_one_week = generate_table_html(one_week_dates[::-1])
    table_html_three_month = generate_table_html(three_month_dates[::-1])
    table_html_all = generate_table_html(all_dates[::-1]) if full_mode else table_html_three_month

    # 生成板块涨跌统计表格行（只显示TOP20）
    sector_change_stats_sorted = sorted(sector_change_stats, key=lambda x: x['limit_up_count'], reverse=True)
    
    sector_change_rows = []
    for stat in sector_change_stats_sorted[:20]:
        # 判断该板块是否在最新日期的TOP20中
        is_in_top20 = stat['sector_code'] in top20_sector_codes
        
        # 根据是否在TOP20中决定行底色和字体颜色
        row_class = ""
        sector_name_class = ""
        if is_in_top20:
            # 在TOP20中 - 浅红色背景，红色字体
            row_class = ' row-positive'
            sector_name_class = ' sector-name-positive'

        # 涨幅比显示值 - 所有板块按正负值着色，虚拟板块板块涨幅旁标注「主题」
        if stat.get('is_virtual'):
            avg_satisfied = stat['avg_satisfied_change']
            avg_all = stat['avg_all_change']
            sat_color = '#d32f2f' if avg_satisfied > 0 else ('#388e3c' if avg_satisfied < 0 else '#666')
            all_color = '#d32f2f' if avg_all > 0 else ('#388e3c' if avg_all < 0 else '#666')
            change_display = (
                f"<span style='color:{sat_color};font-weight:bold;'>{avg_satisfied:.2f}%</span>"
                f" / <span style='color:{all_color};font-weight:bold;'>{avg_all:.2f}%</span>"
            )
        else:
            # 满足股票平均涨幅颜色
            if stat['avg_satisfied_change'] > 0:
                satisfied_change_class = ' change-positive'
            elif stat['avg_satisfied_change'] < 0:
                satisfied_change_class = ' change-negative'
            else:
                satisfied_change_class = ''
            
            # 板块平均涨幅颜色
            if stat['avg_all_change'] > 0:
                all_change_class = ' change-positive'
            elif stat['avg_all_change'] < 0:
                all_change_class = ' change-negative'
            else:
                all_change_class = ''

            change_display = f"<span class='{satisfied_change_class}'>{stat['avg_satisfied_change']:.2f}%</span> / <span class='{all_change_class}'>{stat['avg_all_change']:.2f}%</span>"

        # 涨停数样式（大于0时加粗加红）
        limit_up_count_class = " limit-up-count" if stat['limit_up_count'] > 0 else ""

        row = f'''<tr data-sector="{stat['sector_code']}" data-limit-up="{stat['limit_up_ratio']}" data-limit-up-count="{stat['limit_up_count']}" data-change="{stat['change_ratio']}" data-sector-change="{stat['avg_all_change']}" data-is-virtual="{'1' if stat.get('is_virtual') else '0'}" class="{row_class}">
                <td class="{sector_name_class}" data-sector-code="{stat['sector_code']}"><div class="sector-name" onclick="highlightSector('{stat['sector_code']}')" ondblclick="showSectorTrend('{stat['sector_code']}', '{stat['sector_name']}')">{stat['sector_name']}</div></td>
                <td><span class="{limit_up_count_class}">{stat['limit_up_count']}</span>/{stat['analyzed_count']}<br><span style="color:#666;font-size:0.9em;">{stat['limit_up_ratio']:.1f}%</span></td>
                <td>{change_display}</td>
              </tr>'''
        sector_change_rows.append(row)
    sector_change_rows = '\n'.join(sector_change_rows) if sector_change_rows else '<tr><td colspan="3" style="text-align:center;color:#bbb;padding:20px;">暂无数据</td></tr>'

    # ---- 加载所有板块的历史趋势数据 ----
    # 获取所有出现过的板块代码（包括右侧趋势跟踪表格和左侧板块涨跌统计表格）
    all_sector_codes = set()
    # 从右侧趋势跟踪表格获取板块代码
    for date_str in all_dates:
        daily_data = daily_top20.get(date_str, {})
        for sector in daily_data.get('sectors', []):
            all_sector_codes.add(sector['code'])
    # 从左侧板块涨跌统计表格获取板块代码
    for stat in sector_change_stats:
        all_sector_codes.add(stat['sector_code'])

    # 加载每个板块的历史数据（限制为最近三个月）
    sector_trend_data = {}
    for code in all_sector_codes:
        dates, counts, rates = load_sector_trend(conn, code, period=period)
        # 根据 full_mode 切片数据
        if full_mode:
            sliced_dates = dates
            sliced_counts = counts
            sliced_rates = rates
        else:
            cutoff = max(0, len(dates) - 60)
            sliced_dates = dates[cutoff:]
            sliced_counts = counts[cutoff:]
            sliced_rates = rates[cutoff:]
        sector_trend_data[code] = {
            'dates': sliced_dates,
            'counts': sliced_counts,
            'rates': sliced_rates
        }

    # 将板块数据转换为JSON字符串用于嵌入HTML
    sector_trend_json = json.dumps(sector_trend_data, ensure_ascii=False)
    # 注意：三周期数据(_sectorMultiPeriodData)和弹窗逻辑在页面底部只定义一次

    # ---- 总体满足数趋势（Chart.js 混合图表）----
    # 注意：使用 global_dates（来自 t_run_log）作为实际图表日期数量
    # 而非 all_dates（来自 t_daily_report 的 TOP20 表格）
    global_dates, global_sat_counts, global_sat_rates = load_global_trend(conn, period=period)
    trend_hint_date_count = len(global_dates) if global_dates else len(all_dates)

    # 全局趋势图数据也限制为最近三个月
    if full_mode:
        chart_dates = global_dates
        chart_sat_counts = global_sat_counts
        chart_sat_rates = global_sat_rates
    else:
        cutoff = max(0, len(global_dates) - 60)
        chart_dates = global_dates[cutoff:]
        chart_sat_counts = global_sat_counts[cutoff:]
        chart_sat_rates = global_sat_rates[cutoff:]

    chart_html = ""
    if chart_dates and len(chart_dates) > 1:
        chart_html = f'''
    <div class="trend-chart-section">
      <div class="chart-date-filter" style="margin-bottom:4px;">
        <h3 class="trend-chart-title" style="margin:0;font-size:1em;">市场总满足数趋势图：</h3>
        <button class="chart-filter-btn" data-range="one_month" onclick="updateChartRange_{period}('one_month')">一个月</button>
        <button class="chart-filter-btn active" data-range="three_month" onclick="updateChartRange_{period}('three_month')">三个月</button>
        {"""<button class="chart-filter-btn" data-range="one_year" onclick="updateChartRange_{period}('one_year')">一年</button>
        <button class="chart-filter-btn" data-range="all" onclick="updateChartRange_{period}('all')">全部</button>""" if full_mode else ""}
      </div>
      <div style="height: 250px; width: 100%;">
        <canvas id="globalTrendChart_{period}"></canvas>
      </div>
    </div>
    <script>
      // Tab 专属数据（每个 Tab 独立作用域）
      var _fullDates_{period} = {json.dumps(chart_dates)};
      var _fullSatCounts_{period} = {json.dumps(chart_sat_counts)};
      var _fullSatRates_{period} = {json.dumps(chart_sat_rates)};
      var _trendChart_{period} = null;

      // 初始化该 Tab 的图表和表格（在 DOMContentLoaded 时以及 Tab 切换时调用）
      function initChartForTab_{period}() {{
        currentTabPeriod = '{period}';
        // 检查 filterTrendData 函数是否已定义
        if (typeof window['filterTrendData_{period}'] === 'function') {{
          filterTrendData_{period}('two_week');
        }} else {{
          console.warn('filterTrendData_{period} 函数未定义，跳过表格初始化');
        }}
        updateChartRange_{period}('three_month');
      }}

      // 如果该 Tab 当前可见，立即初始化（仅执行一次）
      if (document.getElementById('tab-{period}').classList.contains('active')) {{
        window.addEventListener('DOMContentLoaded', function() {{
          initChartForTab_{period}();
        }});
      }}

      function updateChartRange_{period}(range) {{
        var buttons = document.querySelectorAll('#tab-{period} .chart-filter-btn');
        buttons.forEach(btn => btn.classList.remove('active'));
        for (var i = 0; i < buttons.length; i++) {{
          if (buttons[i].getAttribute('data-range') === range) {{
            buttons[i].classList.add('active');
            break;
          }}
        }}

        var startIndex = 0;
        if (range === 'one_month') {{
          startIndex = Math.max(0, _fullDates_{period}.length - 20);
        }} else if (range === 'three_month') {{
          startIndex = Math.max(0, _fullDates_{period}.length - 60);
        }} else if (range === 'one_year') {{
          startIndex = Math.max(0, _fullDates_{period}.length - 240);
        }}

        var displayDates = _fullDates_{period}.slice(startIndex);
        var displaySatCounts = _fullSatCounts_{period}.slice(startIndex);
        var displaySatRates = _fullSatRates_{period}.slice(startIndex);

        if (_trendChart_{period}) {{
          _trendChart_{period}.destroy();
        }}

        var ctx = document.getElementById('globalTrendChart_{period}');
        _trendChart_{period} = new Chart(ctx, {{
          type: 'bar',
          data: {{
            labels: displayDates,
            datasets: [
              {{
                label: '满足数',
                data: displaySatCounts,
                type: 'bar',
                yAxisID: 'y-count-{period}',
                backgroundColor: 'rgba(102, 126, 234, 0.7)',
                borderColor: 'rgba(102, 126, 234, 1)',
                borderWidth: 1,
                borderRadius: 4
              }},
              {{
                label: '满足率(%)',
                data: displaySatRates,
                type: 'line',
                yAxisID: 'y-rate-{period}',
                borderColor: '#f5576c',
                backgroundColor: '#f5576c',
                tension: 0.3,
                pointRadius: 3,
                pointHoverRadius: 5
              }}
            ]
          }},
          options: {{
            responsive: true,
            maintainAspectRatio: false,
            interaction: {{
              mode: 'index',
              intersect: false
            }},
            plugins: {{
              legend: {{
                position: 'top',
                labels: {{
                  usePointStyle: true,
                  padding: 15,
                  font: {{ size: 12 }}
                }}
              }},
              tooltip: {{
                backgroundColor: 'rgba(0, 0, 0, 0.8)',
                titleFont: {{ size: 13 }},
                bodyFont: {{ size: 12 }},
                padding: 10,
                cornerRadius: 6,
                callbacks: {{
                  label: function(context) {{
                    var label = context.dataset.label || '';
                    if (label) {{
                      label += ': ';
                    }}
                    if (context.dataset.type === 'line') {{
                      label += context.parsed.y.toFixed(2) + '%';
                    }} else {{
                      label += context.parsed.y + ' 只';
                    }}
                    return label;
                  }}
                }}
              }}
            }},
            scales: {{
              x: {{
                grid: {{
                  display: false
                }},
                ticks: {{
                  font: {{ size: 10 }},
                  maxRotation: 30,
                  minRotation: 30
                }}
              }},
              'y-count-{period}': {{
                type: 'linear',
                position: 'left',
                title: {{
                  display: true,
                  text: '满足数',
                  font: {{ size: 11 }}
                }},
                grid: {{
                  color: 'rgba(0, 0, 0, 0.05)'
                }},
                ticks: {{
                  font: {{ size: 11 }}
                }}
              }},
              'y-rate-{period}': {{
                type: 'linear',
                position: 'right',
                title: {{
                  display: true,
                  text: '满足率(%)',
                  font: {{ size: 11 }}
                }},
                grid: {{
                  display: false
                }},
                ticks: {{
                  font: {{ size: 11 }}
                }}
              }}
            }}
          }}
        }});
      }}
    </script>
        '''
    elif global_dates and len(global_dates) == 1:
        chart_html = '<div class="no-trend">当前仅有 1 天的历史数据，继续运行后将显示趋势图</div>'
    else:
        chart_html = '<div class="no-trend">暂无历史数据，运行回溯脚本后将显示趋势图</div>'

    # full_mode 时的附加 JS 数据
    if full_mode:
        all_entry = ",\n      'all': `" + table_html_all.replace('`', '\\`') + "`"
        btn_all = ", 'all'"
    else:
        all_entry = ""
        btn_all = ""

    return f"""
    <div class="trend-hint">
      总共 {trend_hint_date_count} 天数据 | 每天独立TOP20板块 |
      <span class="trend-up">▲ 上升</span>
      <span class="trend-down">▼ 下降</span>（变动 &gt; 0.5% 时标色）
    </div>
    {chart_html}

    <!-- 注意：板块趋势弹窗(sectorModal)在页面底部只定义一次 -->

    <!-- 下方分栏：左栏预留表格 + 右栏TOP20表格 -->
    <div class="trend-bottom-columns">

      <!-- 左栏：板块涨跌统计表格 -->
      <div class="trend-left-col">
        <div class="trend-chart-section" style="height:100%;">
          <h3 class="trend-chart-title">板块涨跌统计</h3>
          <table class="reserve-table" id="sectorChangeTable_{period}">
            <thead>
              <tr>
                <th onclick="sortSectorTable_{period}('sector_change')" style="cursor:pointer;">板块 <span id="sort-sector_change"></span></th>
                <th onclick="sortSectorTable_{period}('limit_up_count')" style="cursor:pointer;">涨停比 <span id="sort-limit_up_count">↓</span></th>
                <th onclick="sortSectorTable_{period}('sector_change')" style="cursor:pointer;">涨幅比 <span id="sort-sector_change2"></span></th>
              </tr>
            </thead>
            <tbody id="sectorChangeBody_{period}">
{sector_change_rows}
            </tbody>
          </table>
        </div>
      </div>

      <!-- 右栏：日期筛选 + TOP20表格 -->
      <div class="trend-right-col">
        <div class="trend-chart-section">
          <div class="date-filter-bar">
            <span class="filter-label">日期筛选：</span>
            <button class="filter-btn active" onclick="filterTrendData_{period}('two_week')">最近二周</button>
            <button class="filter-btn" onclick="filterTrendData_{period}('one_month')">最近一个月</button>
            <button class="filter-btn" onclick="filterTrendData_{period}('three_month')">最近三个月</button>
            {"<button class=\"filter-btn\" onclick=\"filterTrendData('all')\">全部数据</button>" if full_mode else ""}
            <button class="export-btn" onclick="exportTrendData_{period}()" style="margin-left: auto; background: #1cb85e; color: white; border: none; padding: 6px 16px; border-radius: 4px; cursor: pointer; font-size: 14px;">导出当前数据</button>
          </div>
          <h3 class="trend-chart-title">每日TOP20板块总比满足率变化（表格）</h3>
          <div class="trend-scroll">
            <table class="trend-table" id="trendTable_{period}">
              <thead id="trendTableHead_{period}"></thead>
              <tbody id="trendTableBody_{period}"></tbody>
            </table>
          </div>
        </div>
      </div>

    </div>

    <script>
    // 存储不同日期范围的表格HTML
    var _trendTableData_{period} = {{
      'two_week': `{{{table_html_two_week}}}`,
      'one_month': `{{{table_html_one_month}}}`,
      'three_month': `{{{table_html_three_month}}}`{all_entry}
    }};

    // 当前选中的时间范围
    var _currentRange_{period} = 'two_week';

    // 默认显示最近二周
    // (called from initChartForTab)

    function filterTrendData_{period}(range) {{
      // 记录当前选中的时间范围
      _currentRange_{period} = range;

      // 根据range值更新对应的按钮状态
      var buttons = document.querySelectorAll('#tab-{period} .filter-btn');
      var buttonTexts = ['two_week', 'one_month', 'three_month'{btn_all}];
      var idx = buttonTexts.indexOf(range);

      if (idx >= 0 && buttons[idx]) {{
        // 移除所有active类,然后添加到对应按钮
        buttons.forEach(btn => btn.classList.remove('active'));
        buttons[idx].classList.add('active');
      }}

      // 获取对应的表格HTML
      var tableHtml = _trendTableData_{period}[range];
      if (!tableHtml) return;

      // 解析HTML并更新表格
      var thead = document.getElementById('trendTableHead_{period}');
      var tbody = document.getElementById('trendTableBody_{period}');

      // 创建临时容器解析HTML
      var tempDiv = document.createElement('div');
      tempDiv.innerHTML = `<table>${'{tableHtml}'}</table>`;

      // 更新表头和tbody
      var tempTable = tempDiv.querySelector('table');
      var tempThead = tempTable.querySelector('thead');
      var tempTbody = tempTable.querySelector('tbody');

      if (tempThead) {{
        thead.innerHTML = tempThead.innerHTML;
      }}
      if (tempTbody) {{
        tbody.innerHTML = tempTbody.innerHTML;
      }}

      // 清除所有高亮
      document.querySelectorAll('.sector-highlighted').forEach(el => {{
        el.classList.remove('sector-highlighted');
      }});
    }}

    // 存储所有板块的趋势数据
    var _sectorTrendData_{period} = {sector_trend_json};
    // 注意：_sectorMultiPeriodData 在页面底部只定义一次（共享）

    // 导出趋势表格数据为Excel
    function exportTrendData_{period}() {{
      var tableHtml = _trendTableData_{period}[_currentRange_{period}];
      if (!tableHtml) {{
        alert('没有可导出的数据');
        return;
      }}

      // 创建临时表格解析HTML
      var tempDiv = document.createElement('div');
      tempDiv.innerHTML = '<table>' + tableHtml + '</table>';
      var tempTable = tempDiv.querySelector('table');
      var rows = tempTable2.querySelectorAll('tr');

      if (rows.length === 0) {{
        alert('表格数据为空');
        return;
      }}

      // 构建CSV内容
      let csvContent = '\uFEFF'; // BOM for Excel UTF-8

      rows.forEach(row => {{
        const cells = row.querySelectorAll('th, td');
        const rowData = [];
        cells.forEach(cell => {{
          // 提取纯文本，移除HTML标签
          let text = cell.textContent || cell.innerText || '';
          // 处理CSV特殊字符
          text = text.replace(/"/g, '""');
          if (text.includes(',') || text.includes('"') || text.includes(String.fromCharCode(10))) {{
            text = '"' + text + '"';
          }}
          rowData.push(text);
        }});
        csvContent += rowData.join(',') + String.fromCharCode(10);
      }});

      // 生成文件名
      const dateStr = new Date().toISOString().slice(0, 10).replace(/-/g, '');
      const periodNames = {{
        'two_week': '最近二周',
        'one_month': '最近一个月',
        'three_month': '最近三个月',
        'all': '全部数据'
      }};
      var filename = '趋势跟踪_' + (periodNames[currentPeriod] || currentPeriod) + '_' + dateStr + '.csv';

      // 下载文件
      const blob = new Blob([csvContent], {{ type: 'text/csv;charset=utf-8;' }});
      const link = document.createElement('a');
      link.href = URL.createObjectURL(blob);
      link.download = filename;
      link.style.display = 'none';
      document.body.appendChild(link);
      link.click();
      document.body.removeChild(link);

      console.log('已导出:', filename);
    }}

    // 板块涨跌统计表格排序
    var _currentSort_{period} = {{ field: 'limit_up_count', order: 'desc' }};  // 默认按涨停数量降序

    function sortSectorTable_{period}(field) {{
      var tbody = document.getElementById('sectorChangeBody_{period}');
      if (!tbody) return;

      const rows = Array.from(tbody.querySelectorAll('tr'));
      if (rows.length === 0) return;

      // 切换排序顺序
      if (_currentSort_{period}.field === field) {{
        _currentSort_{period}.order = _currentSort_{period}.order === 'desc' ? 'asc' : 'desc';
      }} else {{
        _currentSort_{period}.field = field;
        _currentSort_{period}.order = 'desc';  // 默认降序
      }}

      // 更新排序指示器
      document.getElementById('sort-sector_change_{period}').textContent = '';
      document.getElementById('sort-limit_up_count_{period}').textContent = '';
      document.getElementById('sort-sector_change2_{period}').textContent = '';
      const indicator = currentSort.order === 'desc' ? '↓' : '↑';
      
      // 根据字段更新对应的指示器
      if (field === 'limit_up_count') {{
        document.getElementById('sort-limit_up_count_{period}').textContent = indicator;
      }} else if (field === 'sector_change') {{
        // 涨幅比列按板块涨幅排序
        document.getElementById('sort-sector_change2_{period}').textContent = indicator;
      }}

      // 排序行（虚拟板块始终排在末尾）
      rows.sort((a, b) => {{
        // 虚拟板块固定排末尾
        const aIsVirtual = a.dataset.isVirtual === '1';
        const bIsVirtual = b.dataset.isVirtual === '1';
        if (aIsVirtual && !bIsVirtual) return 1;
        if (!aIsVirtual && bIsVirtual) return -1;

        let valA, valB;
        if (field === 'limit_up_count') {{
          // 涨停比列按涨停数量排序
          valA = parseInt(a.dataset.limitUpCount) || 0;
          valB = parseInt(b.dataset.limitUpCount) || 0;
        }} else if (field === 'sector_change') {{
          // 涨幅比列按板块涨幅排序
          valA = parseFloat(a.dataset.sectorChange) || 0;
          valB = parseFloat(b.dataset.sectorChange) || 0;
        }} else {{
          valA = parseFloat(a.dataset.change) || 0;
          valB = parseFloat(b.dataset.change) || 0;
        }}
        return _currentSort_{period}.order === 'desc' ? valB - valA : valA - valB;
      }});

      // 重新插入排序后的行
      rows.forEach(row => tbody.appendChild(row));
    }}

    // 板块趋势图表实例（单周期）
    let sectorTrendChart_{period} = null;
    // 注意：sectorMultiPeriodChart 和弹窗逻辑在页面底部只定义一次

    // 显示板块趋势弹窗
    function showSectorTrend(sectorCode, sectorName) {{
      const modal = document.getElementById('sectorModal');
      const title = document.getElementById('modalTitle');

      // 设置标题
      title.textContent = `板块趋势：${{sectorName}}（满足数 vs 总比满足率）`;

      // 显示弹窗
      modal.style.display = 'block';

      // 获取该板块的数据（原有单周期数据）
      var data = _sectorTrendData_{period}[sectorCode];
      if (!data) {{
        console.error('板块数据未找到:', sectorCode);
        return;
      }}

      // ---- 初始化三周期满足数趋势图（由共享函数处理）----
      if (typeof initMultiPeriodChart === 'function') {{
        initMultiPeriodChart(sectorCode);
      }}

      // 如果已存在图表，销毁它
      if (sectorTrendChart_{period}) {{
        sectorTrendChart_{period}.destroy();
      }}

      // 创建新的图表
      const ctx = document.getElementById('sectorTrendChart');
      sectorTrendChart_{period} = new Chart(ctx, {{
        type: 'bar',
        data: {{
          labels: data.dates,
          datasets: [
            {{
              label: '满足数',
              data: data.counts,
              type: 'bar',
              yAxisID: 'y-count',
              backgroundColor: 'rgba(102, 126, 234, 0.7)',
              borderColor: 'rgba(102, 126, 234, 1)',
              borderWidth: 1,
              borderRadius: 4
            }},
            {{
              label: '总比满足率(%)',
              data: data.rates,
              type: 'line',
              yAxisID: 'y-rate',
              borderColor: '#f5576c',
              backgroundColor: '#f5576c',
              tension: 0.3,
              pointRadius: 3,
              pointHoverRadius: 5
            }}
          ]
        }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          interaction: {{
            mode: 'index',
            intersect: false
          }},
          plugins: {{
            legend: {{
              position: 'top',
              labels: {{
                usePointStyle: true,
                padding: 15,
                font: {{ size: 12 }}
              }}
            }},
            tooltip: {{
              backgroundColor: 'rgba(0, 0, 0, 0.8)',
              titleFont: {{ size: 13 }},
              bodyFont: {{ size: 12 }},
              padding: 10,
              cornerRadius: 6,
              callbacks: {{
                label: function(context) {{
                  var label = context.dataset.label || '';
                  if (label) {{
                    label += ': ';
                  }}
                  if (context.dataset.type === 'line') {{
                    label += context.parsed.y.toFixed(2) + '%';
                  }} else {{
                    label += context.parsed.y + ' 只';
                  }}
                  return label;
                }}
              }}
            }}
          }},
          scales: {{
            x: {{
              grid: {{
                display: false
              }},
              ticks: {{
                font: {{ size: 10 }},
                maxRotation: 30,
                minRotation: 30
              }}
            }},
            'y-count': {{
              type: 'linear',
              position: 'left',
              title: {{
                display: true,
                text: '满足数',
                font: {{ size: 11 }}
              }},
              grid: {{
                color: 'rgba(0, 0, 0, 0.05)'
              }},
              ticks: {{
                font: {{ size: 11 }}
              }}
            }},
            'y-rate': {{
              type: 'linear',
              position: 'right',
              title: {{
                display: true,
                text: '总比满足率(%)',
                font: {{ size: 11 }}
              }},
              grid: {{
                display: false
              }},
              ticks: {{
                font: {{ size: 11 }}
              }}
            }}
          }}
        }}
      }});
    }}

    // 关闭板块趋势弹窗
    function closeSectorModal() {{
      const modal = document.getElementById('sectorModal');
      modal.style.display = 'none';

      // 销毁图表以释放资源
      if (sectorTrendChart_{period}) {{
        sectorTrendChart_{period}.destroy();
        sectorTrendChart_{period} = null;
      }}
      // 三周期图表由共享函数销毁
      if (typeof destroyMultiPeriodChart === 'function') {{
        destroyMultiPeriodChart();
      }}
    }}

    // 注意：window.onclick 关闭弹窗逻辑在页面底部只定义一次
    </script>"""


def build_shared_modal_js(conn):
    """生成只出现一次的共享部分：弹窗HTML + 三周期数据 + 三周期图表逻辑
    
    这些内容被三个Tab的showSectorTrend/closeSectorModal共同引用，
    不能放在 build_trend_html() 中（因为该函数被调用3次会导致重复声明）。
    """
    # 获取所有板块代码
    all_sector_codes_row = conn.execute("SELECT sector_code FROM t_sector").fetchall()
    all_sector_codes = [r[0] for r in all_sector_codes_row]
    
    # 加载三周期数据
    sector_multi_period_data = {}
    for code in all_sector_codes:
        mp_data = load_sector_multi_period_trend(conn, code)
        sector_multi_period_data[code] = mp_data
    sector_multi_period_json = json.dumps(sector_multi_period_data, ensure_ascii=False)
    
    return f"""
    <!-- 板块趋势弹窗（全局唯一，三个Tab共享） -->
    <div id="sectorModal" class="modal" style="display:none;">
      <div class="modal-content">
        <div class="modal-header">
          <h3 id="modalTitle">板块趋势</h3>
          <span class="modal-close" onclick="closeSectorModal()">&times;</span>
        </div>
        <div class="modal-body">
          <!-- 三周期满足数趋势图 -->
          <div style="margin-bottom:16px;">
            <div style="font-size:0.9em;color:#555;margin-bottom:6px;font-weight:bold;">三周期满足数趋势</div>
            <div style="height: 280px; width: 100%;">
              <canvas id="sectorMultiPeriodChart"></canvas>
            </div>
          </div>
          <!-- 当前周期趋势图 -->
          <div style="height: 350px; width: 100%;">
            <canvas id="sectorTrendChart"></canvas>
          </div>
        </div>
      </div>
    </div>

    <script>
      // 存储所有板块的三周期满足数趋势数据（全局唯一）
      var _sectorMultiPeriodData = {sector_multi_period_json};

      // 三周期趋势图实例（全局唯一，避免重复声明）
      let sectorMultiPeriodChart = null;

      // 初始化三周期满足数趋势图（由各Tab的showSectorTrend调用）
      function initMultiPeriodChart(sectorCode) {{
        var mpData = _sectorMultiPeriodData[sectorCode];
        if (!mpData || !mpData.dates || mpData.dates.length === 0) {{
          console.warn('无三周期数据:', sectorCode);
          return;
        }}

        // 销毁旧实例
        if (sectorMultiPeriodChart) {{
          sectorMultiPeriodChart.destroy();
        }}
        var mpCtx = document.getElementById('sectorMultiPeriodChart');
        if (!mpCtx) return;

        sectorMultiPeriodChart = new Chart(mpCtx, {{
          type: 'line',
          data: {{
            labels: mpData.dates,
            datasets: [
              {{
                label: '日线满足数',
                data: mpData.daily,
                borderColor: '#ff0000',
                backgroundColor: '#ff0000',
                tension: 0.3,
                pointRadius: 1,
                pointHoverRadius: 4,
                borderWidth: 2
              }},
              {{
                label: '三日满足数',
                data: mpData['3day'],
                borderColor: '#000080',
                backgroundColor: '#000080',
                tension: 0.3,
                pointRadius: 1,
                pointHoverRadius: 4,
                borderWidth: 2
              }},
              {{
                label: '周线满足数/4',
                data: mpData.weekly,
                borderColor: '#ffff00',
                backgroundColor: '#ffff00',
                tension: 0.3,
                pointRadius: 1,
                pointHoverRadius: 4,
                borderWidth: 2
              }}
            ]
          }},
          options: {{
            responsive: true,
            maintainAspectRatio: false,
            interaction: {{
              mode: 'index',
              intersect: false
            }},
            plugins: {{
              legend: {{
                position: 'top',
                labels: {{
                  usePointStyle: true,
                  padding: 12,
                  font: {{ size: 11 }}
                }}
              }},
              tooltip: {{
                backgroundColor: 'rgba(0, 0, 0, 0.8)',
                titleFont: {{ size: 12 }},
                bodyFont: {{ size: 11 }},
                padding: 8,
                cornerRadius: 6,
                callbacks: {{
                  label: function(context) {{
                    var label = context.dataset.label || '';
                    if (label) label += ': ';
                    var val = context.parsed.y;
                    if (val !== null && val !== undefined) {{
                      label += parseFloat(val).toFixed(2);
                    }} else {{
                      label += '-';
                    }}
                    return label;
                  }}
                }}
              }}
            }},
            scales: {{
              x: {{
                grid: {{ display: false }},
                ticks: {{ font: {{ size: 9 }}, maxRotation: 30, minRotation: 30 }}
              }},
              y: {{
                type: 'linear',
                position: 'left',
                title: {{ display: true, text: '满足数', font: {{ size: 10 }} }},
                grid: {{ color: 'rgba(0,0,0,0.05)' }},
                ticks: {{ font: {{ size: 10 }} }}
              }}
            }}
          }}
        }});
      }}

      // 销毁三周期图表（由各Tab的closeSectorModal调用）
      function destroyMultiPeriodChart() {{
        if (sectorMultiPeriodChart) {{
          sectorMultiPeriodChart.destroy();
          sectorMultiPeriodChart = null;
        }}
      }}

      // 点击弹窗外部关闭
      window.onclick = function(event) {{
        const modal = document.getElementById('sectorModal');
        if (event.target == modal) {{
          closeSectorModal();
        }}
      }};
    </script>"""


def export_stocks_to_excel(conn, run_id, all_sectors, output_dir):
    """导出满足条件个股明细到Excel"""
    # 从run_id提取日期 (格式: 20260325_HIST -> 20260325)
    report_date = run_id.split('_')[0] if '_' in run_id else datetime.now().strftime('%Y%m%d')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "满足条件个股明细"

    # 表头样式
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    headers = ["板块排名", "板块名称", "板块满足率", "股票代码", "股票名称", "妖线角度", "主升1", "收盘价"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = header_alignment

    # 写入数据
    row = 2
    for item in all_sectors[:20]:
        if item['satisfied'] <= 0:
            continue
        stocks = load_satisfied_stocks(conn, run_id, item['sector_code'], top_n=30)
        if not stocks:
            continue
        for st in stocks:
            ws.cell(row=row, column=1, value=item['rank'])
            ws.cell(row=row, column=2, value=item['name'])
            ws.cell(row=row, column=3, value=f"{item['rate']:.2f}%")
            ws.cell(row=row, column=4, value=st['code'])
            ws.cell(row=row, column=5, value=st['name'])
            ws.cell(row=row, column=6, value=f"{st['angle']:.2f}")
            ws.cell(row=row, column=7, value=f"{st['zhusheng1']:.4f}" if st['zhusheng1'] else "-")
            ws.cell(row=row, column=8, value=f"{st['close']:.2f}" if st['close'] else "-")
            row += 1

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # 保存文件（以日期命名，默认覆盖）
    filename = f"{output_dir}/satisfied_stocks_{report_date}.xlsx"
    wb.save(filename)
    print(f"[导出] 个股明细已导出: {filename}")
    return filename


def export_trend_to_excel(conn, all_dates, daily_top20, output_dir, run_id=None):
    """导出趋势跟踪表格到Excel"""
    # 从run_id提取日期，如果没有则使用最新日期
    if run_id and '_' in run_id:
        report_date = run_id.split('_')[0]
    elif all_dates:
        report_date = all_dates[-1].replace('-', '')
    else:
        report_date = datetime.now().strftime('%Y%m%d')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "每日TOP20板块总比满足率变化"

    # 表头样式
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    headers = ["排名"] + [d[5:] for d in all_dates]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = header_alignment

    # 写入总数行
    ws.cell(row=2, column=1, value="总数")
    for col_idx, date_str in enumerate(all_dates, start=2):
        total_count = daily_top20.get(date_str, {}).get('total_satisfied', 0)
        ws.cell(row=2, column=col_idx, value=total_count)
        ws.cell(row=2, column=col_idx).fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

    # 建立映射：{排名: {日期: 板块数据}}
    rank_data = {}
    for date_str in all_dates:
        daily_data = daily_top20.get(date_str, {})
        sectors = daily_data.get('sectors', [])
        for sector in sectors:
            rank = sector['rank']
            if rank not in rank_data:
                rank_data[rank] = {}
            rank_data[rank][date_str] = sector

    # 写入排名行（1-20）
    for rank in range(1, 21):
        ws.cell(row=rank + 2, column=1, value=rank)
        for col_idx, date_str in enumerate(all_dates, start=2):
            sector = rank_data.get(rank, {}).get(date_str)
            if sector:
                value = f"{sector['name']} ({sector['satisfied_count']})\n{sector['total_rate']:.2f}%"
                ws.cell(row=rank + 2, column=col_idx, value=value)
            else:
                ws.cell(row=rank + 2, column=col_idx, value="-")

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # 设置行高
    for row_num in range(3, 23):
        ws.row_dimensions[row_num].height = 40

    # 保存文件（以日期命名，默认覆盖）
    filename = f"{output_dir}/trend_tracking_{report_date}.xlsx"
    wb.save(filename)
    print(f"[导出] 趋势跟踪已导出: {filename}")
    return filename


def generate_html(conn, run_id_daily, run_id_3day, run_id_weekly, output_file, report_dir, full_mode=False, use_intraday=False):
    # 为每个周期获取统计数据
    summary_daily  = load_summary(conn, run_id_daily) if run_id_daily else {}
    summary_3day   = load_summary(conn, run_id_3day) if run_id_3day else {}
    summary_weekly = load_summary(conn, run_id_weekly) if run_id_weekly else {}
    
    # 获取历史趋势数据（最近90天）
    historical_data = load_historical_data(conn, days=90)
    
    # 使用日线的 run_id 作为默认 run_id（用于一些兼容性代码）
    default_run_id = run_id_daily or run_id_3day or run_id_weekly
    default_summary = summary_daily or summary_3day or summary_weekly or {}
    
    all_sectors = load_sector_stats(conn, default_run_id) if default_run_id else []
    top20       = all_sectors[:20]

    # 去重后总满足数，用于计算「总比满足率」
    total_satisfied_dedup = default_summary.get('satisfied_count', 0) or 1  # 防零除

    ver      = default_summary.get('version', 'V1.4升级版')
    run_time = default_summary.get('run_time', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    report_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 报告生成时间
    analyzed_count = default_summary.get('analyzed_count', 0)  # 分析的个股数量
    run_id = default_run_id or ''  # 用于模板中显示

    # ---- TOP20 行 ----
    top20_rows = ''
    for item in top20:
        rank_cls  = {1: 'rank-top1', 2: 'rank-top2', 3: 'rank-top3'}.get(item['rank'], '')
        badge     = '<span class="hot-badge">HOT</span>' if item['rank'] == 1 else ''
        total_rate = item['satisfied'] / total_satisfied_dedup * 100
        top20_rows += f"""
          <tr>
            <td class="rank {rank_cls}">{item['rank']}</td>
            <td>{item['name']}{badge}</td>
            <td>{item['total']}</td>
            <td>{item['analyzed']}</td>
            <td>{item['satisfied']}</td>
            <td class="total-rate-val">{total_rate:.2f}%</td>
            <td>
              {item['rate']:.2f}%
              <div class="rate-bar"><div class="rate-fill" style="width:{min(item['rate'],100)}%"></div></div>
            </td>
          </tr>"""

    # ---- 全部板块行 ----
    all_rows = ''
    for item in all_sectors:
        rank_cls   = {1: 'rank-top1', 2: 'rank-top2', 3: 'rank-top3'}.get(item['rank'], '')
        total_rate = item['satisfied'] / total_satisfied_dedup * 100
        all_rows += f"""
          <tr>
            <td class="rank {rank_cls}">{item['rank']}</td>
            <td>{item['name']}</td>
            <td>{item['total']}</td>
            <td>{item['analyzed']}</td>
            <td>{item['satisfied']}</td>
            <td class="total-rate-val">{total_rate:.2f}%</td>
            <td>
              {item['rate']:.2f}%
              <div class="rate-bar"><div class="rate-fill" style="width:{min(item['rate'],100)}%"></div></div>
            </td>
          </tr>"""

    # ---- 满足条件个股明细（TOP 20 板块）----
    stock_detail_html = ''
    for item in all_sectors[:20]:
        if item['satisfied'] <= 0:
            continue
        stocks = load_satisfied_stocks(conn, default_run_id, item['sector_code'], top_n=30)
        if not stocks:
            continue
        rows_html = ''
        for st in stocks:
            zs1_str   = f"{st['zhusheng1']:.4f}" if st['zhusheng1'] is not None else '-'
            close_str = f"{st['close']:.2f}"     if st['close']     is not None else '-'
            rows_html += f"""
              <tr>
                <td>{st['code']}</td>
                <td class="stock-name">{st['name']}</td>
                <td class="angle-val">{st['angle']:.2f}</td>
                <td>{zs1_str}</td>
                <td>{close_str}</td>
              </tr>"""
        stock_detail_html += f"""
        <div class="stock-group">
          <div class="stock-group-title">
            [{item['rank']}] {item['name']}
            <span class="rate-badge">{item['rate']:.2f}%</span>
            <span class="count-badge">满足 {item['satisfied']} / 分析 {item['analyzed']}</span>
          </div>
          <table class="stock-table">
            <thead>
              <tr><th>代码</th><th>名称</th><th>妖线角度</th><th>主升1</th><th>收盘价</th></tr>
            </thead>
            <tbody>{rows_html}</tbody>
          </table>
        </div>"""

    # ---- 趋势跟踪区块（三个周期各自生成）----
    trend_html_daily  = build_trend_html(conn, period='daily',  full_mode=full_mode, use_intraday=use_intraday)
    trend_html_3day   = build_trend_html(conn, period='3day',   full_mode=full_mode, use_intraday=use_intraday)
    trend_html_weekly = build_trend_html(conn, period='weekly', full_mode=full_mode, use_intraday=use_intraday)

    # ---- 共享弹窗和三周期逻辑（只生成一次）----
    shared_modal_js = build_shared_modal_js(conn)

    # Chart.js 文件路径 - 使用相对路径
    chart_js_path = f"./chart.umd.min.js"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <title>概念板块周线妖线角度分析报告 - {ver}</title>
  <script src="{chart_js_path}"></script>
  <style>
    *{{margin:0;padding:0;box-sizing:border-box}}
    body{{font-family:'Microsoft YaHei',Arial,sans-serif;
          background:linear-gradient(135deg,#667eea,#764ba2);
          min-height:100vh;padding:20px}}
    .container{{max-width:1400px;margin:0 auto;background:#fff;
                border-radius:20px;box-shadow:0 20px 60px rgba(0,0,0,.3);overflow:hidden}}

    /* Header */
    .header{{background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;
             padding:24px 40px;text-align:center}}
    .header h1{{font-size:2em;margin-bottom:6px}}
    .header .subtitle{{font-size:.9em;opacity:.85}}
    .version-tag{{display:inline-block;
                  background:linear-gradient(135deg,#f093fb,#f5576c);
                  color:#fff;padding:3px 12px;border-radius:12px;
                  font-size:.82em;font-weight:bold;margin-left:6px}}

    /* DB info bar */
    .db-bar{{background:#1a1b2e;color:#a9dc76;padding:10px 30px;
             font-family:Consolas,monospace;font-size:.8em;
             display:flex;gap:24px;flex-wrap:wrap;align-items:center}}
    .db-bar b{{color:#78dce8}}

    /* Summary cards */
    .summary{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));
              gap:10px;padding:14px 28px;background:#f8f9fa}}
    .card{{background:#fff;padding:10px 16px;border-radius:10px;
           box-shadow:0 2px 10px rgba(0,0,0,.08);text-align:center}}
    .card .val{{font-size:1.5em;font-weight:bold;color:#667eea;margin-bottom:2px}}
    .card .lbl{{color:#888;font-size:.75em}}
    .card.highlight .val{{color:#f5576c}}

    /* Tab Nav */
    .tab-nav{{display:flex;background:#f8f9fa;border-bottom:2px solid #e0e0e0;
              padding:0 28px}}
    .tab{{padding:12px 24px;cursor:pointer;color:#888;font-size:.92em;
          border-bottom:3px solid transparent;margin-bottom:-2px;
          transition:.2s;user-select:none}}
    .tab:hover{{color:#667eea}}
    .tab.active{{color:#667eea;border-bottom-color:#667eea;font-weight:bold}}
    .tab-content{{display:none}}
    .tab-content.active{{display:block}}

    /* Section */
    .section{{padding:28px}}
    .section-title{{font-size:1.5em;color:#333;margin-bottom:16px;
                    padding-bottom:8px;border-bottom:3px solid #667eea}}

    /* 通用表格 */
    table{{width:100%;border-collapse:collapse;margin-top:8px}}
    th,td{{padding:10px 14px;text-align:left;border-bottom:1px solid #eee}}
    th{{background:#f8f9fa;font-weight:bold;color:#333;position:sticky;top:0;z-index:2}}
    tr:hover{{background:#f9f9ff}}
    .rank{{font-weight:bold;color:#667eea}}
    .rank-top1{{color:#f5c518;font-size:1.1em}}
    .rank-top2{{color:#a8a8a8;font-size:1.05em}}
    .rank-top3{{color:#cd7f32;font-size:1.05em}}
    .hot-badge{{background:linear-gradient(135deg,#f093fb,#f5576c);
                color:#fff;padding:2px 8px;border-radius:8px;
                font-size:.75em;font-weight:bold;margin-left:5px}}
    .rate-bar{{width:80px;height:6px;background:#e0e0e0;border-radius:3px;
               overflow:hidden;display:inline-block;vertical-align:middle;margin-left:6px}}
    .rate-fill{{height:100%;background:linear-gradient(90deg,#667eea,#f5576c)}}

    /* 个股明细 */
    .stock-group{{margin-bottom:16px;border:1px solid #e8e8e8;border-radius:8px;overflow:hidden}}
    .stock-group-title{{background:linear-gradient(90deg,#f3f4f8,#eef0f8);
                        padding:9px 16px;font-weight:bold;font-size:.92em;color:#333;
                        border-bottom:1px solid #e8e8e8;display:flex;align-items:center;gap:6px}}
    .rate-badge{{background:#667eea;color:#fff;padding:2px 8px;border-radius:8px;font-size:.76em}}
    .count-badge{{background:#e8f0fe;color:#667eea;padding:2px 8px;border-radius:8px;font-size:.76em}}
    .stock-table th{{background:#fafafa;font-size:.84em;padding:7px 12px}}
    .stock-table td{{font-size:.84em;padding:7px 12px}}
    .stock-name{{color:#555}}
    .angle-val{{font-weight:bold;color:#f5576c}}
    .total-rate-val{{font-weight:bold;color:#2eaf6e}}

    /* 趋势跟踪 */
    .trend-hint{{padding:10px 0 14px;color:#888;font-size:.88em}}
    .trend-hint .trend-up{{color:#e84040;font-weight:bold}}
    .trend-hint .trend-down{{color:#1cb85e;font-weight:bold}}
    /* 图表日期筛选按钮 */
    .chart-date-filter{{padding:12px 0; display:flex; align-items:center; gap:10px; margin-bottom:8px;}}
    .chart-filter-btn{{padding:6px 14px; border:1px solid #ddd; background:#fff; color:#555; border-radius:6px; cursor:pointer; font-size:.82em; transition:all 0.2s; user-select:none;}}
    .chart-filter-btn:hover{{background:#f0f0f0; border-color:#ccc;}}
    .chart-filter-btn.active{{background:linear-gradient(135deg,#667eea,#764ba2); color:#fff; border-color:#667eea; font-weight:bold;}}
    /* 日期筛选按钮 */
    .date-filter-bar{{padding:12px 0; display:flex; align-items:center; gap:10px; margin-bottom:12px;}}
    .filter-label{{font-size:.9em; color:#555; font-weight:bold;}}
    .filter-btn{{padding:8px 16px; border:1px solid #ddd; background:#fff; color:#555; border-radius:6px; cursor:pointer; font-size:.85em; transition:all 0.2s; user-select:none;}}
    .filter-btn:hover{{background:#f0f0f0; border-color:#ccc;}}
    .filter-btn.active{{background:linear-gradient(135deg,#667eea,#764ba2); color:#fff; border-color:#667eea; font-weight:bold;}}
    .trend-scroll{{overflow-x:auto}}
    .trend-table{{min-width:800px;font-size:.84em;border-collapse:collapse}}
    .trend-table th{{background:#f8f9fa;white-space:nowrap;padding:10px 12px;text-align:center;border:1px solid #e0e0e0}}
    .trend-table td{{padding:8px 12px;white-space:nowrap;vertical-align:middle;border:1px solid #e0e0e0}}
    .trend-table .sector-name{{font-weight:bold;color:#333;font-size:.9em;margin-bottom:4px; cursor:pointer; transition:all 0.2s;}}
    .trend-table .sector-name:hover{{color:#667eea; text-decoration:underline;}}
    .trend-table .rate-value{{font-size:.85em;color:#666}}
    .trend-table .rank-cell{{font-weight:bold;color:#667eea;text-align:center;background:#f8f9fa}}
    /* 总数行样式 */
    .trend-table tr:first-child td{{background:#e3f2fd; font-weight:bold; text-align:center; color:#333;}}
    /* 点击高亮：只高亮板块名称部分的字体背景 + 单元格显示框线 */
    .sector-highlighted{{border:2px solid #ffc107 !important;}}
    .sector-highlighted .sector-name{{background:#fff9c4; font-size:1.05em; padding:2px 4px; border-radius:3px;}}
    /* 上升：浅红背景 + 红色文字 + 红色图标 */
    .trend-up{{color:#e84040;background:#fff5f5}}
    .trend-up .sector-name{{color:#e84040}}
    .trend-up .rate-value{{color:#e84040}}
    .trend-up .trend-arrow{{color:#e84040}}
    /* 下降：浅绿背景 + 绿色文字 + 绿色图标 */
    .trend-down{{color:#1cb85e;background:#f5fff8}}
    .trend-down .sector-name{{color:#1cb85e}}
    .trend-down .rate-value{{color:#1cb85e}}
    /* 弹窗样式 */
    .modal{{display:none; position:fixed; z-index:1000; left:0; top:0; width:100%; height:100%; overflow:auto; background-color:rgba(0,0,0,0.5);}}
    .modal-content{{background-color:#fefefe; margin:5% auto; padding:0; border:1px solid #888; width:90%; max-width:900px; border-radius:12px; box-shadow:0 4px 20px rgba(0,0,0,0.3);}}
    .modal-header{{padding:15px 20px; background:linear-gradient(135deg,#667eea,#764ba2); color:#fff; border-radius:12px 12px 0 0; display:flex; justify-content:space-between; align-items:center;}}
    .modal-header h3{{margin:0; font-size:1.3em;}}
    .modal-close{{color:#fff; font-size:28px; font-weight:bold; cursor:pointer; line-height:1; transition:all 0.2s;}}
    .modal-close:hover{{color:#ccc; transform:scale(1.1);}}
    .modal-body{{padding:20px;}}

    .trend-down .trend-arrow{{color:#1cb85e}}

    .trend-chart-section{{margin-top:24px;padding:16px;border:1px solid #e8e8e8;border-radius:8px;background:#fafafa}}
    .trend-chart-title{{font-size:1.1em;color:#333;margin-bottom:12px;font-weight:bold}}
    .no-trend{{padding:30px;text-align:center;color:#aaa;
               background:#f8f9fa;border-radius:8px;font-size:.92em}}

    /* 趋势页下方分栏 */
    .trend-bottom-columns{{display:flex;gap:16px;margin-top:20px;align-items:flex-start}}
    .trend-left-col{{flex:0 0 25%;min-width:0}}
    .trend-right-col{{flex:0 0 75%;min-width:0}}
    /* 左栏板块涨跌统计表格 */
    .reserve-table{{width:100%;border-collapse:collapse;font-size:.82em}}
    .reserve-table th{{background:#f8f9fa;padding:6px 8px;text-align:center;border:1px solid #e0e0e0;font-weight:bold;color:#555;white-space:nowrap}}
    .reserve-table th:hover{{background:#e9ecef}}
    .reserve-table td{{padding:6px 8px;text-align:center;border:1px solid #e0e0e0;font-size:.9em}}
    .reserve-table tr:hover{{background:#f5f5f5}}
    .reserve-table td:first-child{{text-align:left;padding-left:10px}}
    /* 板块涨跌统计 - 正数红色、负数绿色（A股习惯） */
    .reserve-table .change-positive{{color:#c45c5c;font-weight:bold}}
    .reserve-table .change-negative{{color:#1cb85e;font-weight:bold}}
    .reserve-table .limit-up-count{{color:#e74c3c;font-weight:bold}}
    /* 板块名称样式 - 与右侧表格一致，可点击链接样式 */
    .reserve-table .sector-name-positive{{color:#e74c3c;font-weight:bold}}
    .reserve-table .sector-name-negative{{color:#27ae60;font-weight:bold}}
    .reserve-table .sector-name{{cursor:pointer; text-decoration:underline; transition:all 0.2s;}}
    .reserve-table .sector-name:hover{{color:#667eea;}}
    /* 高亮样式 - 左侧和右侧同时高亮 */
    .reserve-table td.sector-highlighted{{border:2px solid #ffc107 !important;}}
    .reserve-table td.sector-highlighted .sector-name{{background:#fff9c4; font-size:1.05em; padding:2px 4px; border-radius:3px;}}
    /* 涨幅比高亮行 - 正数浅红背景，负数浅绿背景 */
    .reserve-table tr.row-positive{{background:#fff5f5}}
    .reserve-table tr.row-positive:hover{{background:#ffe0e0}}
    .reserve-table tr.row-negative{{background:#f5fff8}}
    .reserve-table tr.row-negative:hover{{background:#e0f5e8}}

    /* Footer */
    .footer{{background:#f8f9fa;padding:16px;text-align:center;color:#888;font-size:.82em;line-height:2}}
    .version-badge{{display:inline-block;
                    background:linear-gradient(135deg,#f093fb,#f5576c);
                    color:#fff;padding:3px 12px;border-radius:12px;
                    font-size:.82em;font-weight:bold}}

    @media(max-width:768px){{
      .header h1{{font-size:1.6em}}
      .summary{{grid-template-columns:repeat(2,1fr)}}
      table{{font-size:.82em}}
      .tab{{padding:10px 14px;font-size:.85em}}
    }}
  </style>
</head>
<body>
<div class="container">

  <!-- Header -->
  <div class="header">
    <h1>概念板块周线妖线角度分析</h1>
    <div class="subtitle">268个板块&nbsp;&nbsp;|&nbsp;&nbsp;版本：V1.4升级版&nbsp;&nbsp;|&nbsp;&nbsp;报告时间：{report_time}</div>
  </div>

  <!-- DB信息栏 -->
  <div class="db-bar">
    <span>数据库: <b>{os.path.basename(DB_FILE)}</b></span>
    <span>Run ID: <b>{run_id}</b></span>
    <span>算法: <b>个股计算（{analyzed_count:,}只）</b></span>
    <span>生成: <b>{run_time}</b></span>
  </div>

  <!-- 汇总趋势图 -->
  <div class="section">
    <h2 class="section-title" style="font-size:1.1em;margin-bottom:10px;">三周期满足数趋势汇总</h2>
    <div id="summary-chart" style="width:100%;height:400px;"></div>
  </div>

  <!-- Tab 导航 -->
  <div class="tab-nav">
    <div class="tab active" onclick="switchTab('daily', event)">日线趋势</div>
    <div class="tab" onclick="switchTab('3day', event)">三日趋势</div>
    <div class="tab" onclick="switchTab('weekly', event)">周线趋势</div>
  </div>

  <!-- Tab: 日线趋势 -->
  <div id="tab-daily" class="tab-content active">
    <div class="section">
      <h2 class="section-title" style="font-size:1.1em;margin-bottom:10px;">日线妖线满足率趋势跟踪</h2>
      
      <!-- 日线统计卡片 -->
      <div class="summary" style="margin-bottom:15px;">
        <div class="card">
          <div class="val">{summary_daily.get('sector_count',0)}</div>
          <div class="lbl">分析板块数</div>
        </div>
        <div class="card">
          <div class="val">{summary_daily.get('unique_stocks',0):,}</div>
          <div class="lbl">当前个股</div>
        </div>
        <div class="card">
          <div class="val">{summary_daily.get('analyzed_count',0):,}</div>
          <div class="lbl">成功分析</div>
        </div>
        <div class="card">
          <div class="val">{summary_daily.get('satisfied_count',0):,}</div>
          <div class="lbl">满足条件</div>
        </div>
        <div class="card highlight">
          <div class="val">{summary_daily.get('satisfied_rate',0):.2f}%</div>
          <div class="lbl">整体满足率</div>
        </div>
      </div>
      
      {trend_html_daily}
    </div>
  </div>

  <!-- Tab: 三日趋势 -->
  <div id="tab-3day" class="tab-content">
    <div class="section">
      <h2 class="section-title" style="font-size:1.1em;margin-bottom:10px;">三日妖线满足率趋势跟踪</h2>
      
      <!-- 三日统计卡片 -->
      <div class="summary" style="margin-bottom:15px;">
        <div class="card">
          <div class="val">{summary_3day.get('sector_count',0)}</div>
          <div class="lbl">分析板块数</div>
        </div>
        <div class="card">
          <div class="val">{summary_3day.get('unique_stocks',0):,}</div>
          <div class="lbl">当前个股</div>
        </div>
        <div class="card">
          <div class="val">{summary_3day.get('analyzed_count',0):,}</div>
          <div class="lbl">成功分析</div>
        </div>
        <div class="card">
          <div class="val">{summary_3day.get('satisfied_count',0):,}</div>
          <div class="lbl">满足条件</div>
        </div>
        <div class="card highlight">
          <div class="val">{summary_3day.get('satisfied_rate',0):.2f}%</div>
          <div class="lbl">整体满足率</div>
        </div>
      </div>
      
      {trend_html_3day}
    </div>
  </div>

  <!-- Tab: 周线趋势 -->
  <div id="tab-weekly" class="tab-content">
    <div class="section">
      <h2 class="section-title" style="font-size:1.1em;margin-bottom:10px;">周线妖线满足率趋势跟踪</h2>
      
      <!-- 周线统计卡片 -->
      <div class="summary" style="margin-bottom:15px;">
        <div class="card">
          <div class="val">{summary_weekly.get('sector_count',0)}</div>
          <div class="lbl">分析板块数</div>
        </div>
        <div class="card">
          <div class="val">{summary_weekly.get('unique_stocks',0):,}</div>
          <div class="lbl">当前个股</div>
        </div>
        <div class="card">
          <div class="val">{summary_weekly.get('analyzed_count',0):,}</div>
          <div class="lbl">成功分析</div>
        </div>
        <div class="card">
          <div class="val">{summary_weekly.get('satisfied_count',0):,}</div>
          <div class="lbl">满足条件</div>
        </div>
        <div class="card highlight">
          <div class="val">{summary_weekly.get('satisfied_rate',0):.2f}%</div>
          <div class="lbl">整体满足率</div>
        </div>
      </div>
      
      {trend_html_weekly}
    </div>
  </div>

  <!-- Footer -->
  <div class="footer">
    <p>数据来源：SQLite 数据库 concept_weekly.db | Run ID: {run_id}</p>
    <p>公式: ATAN((主升1/REF(主升1,1)-1)*100)*180/3.1416 &nbsp;|&nbsp; XG: 角度&gt;=63 AND C&gt;=主升1 AND X&gt;0</p>
    <p><span class="version-badge">{ver} 算法优化版</span>
       &nbsp; 报告生成：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
  </div>

  {shared_modal_js}

</div>

  <!-- 引入 ECharts -->
  <script src="./echarts.min.js"></script>

<script>
// 汇总趋势图数据
var summaryChartData = {{
    daily: {json.dumps([{'date': item['date'], 'value': item['value']} for item in historical_data.get('daily', [])])},
    '3day': {json.dumps([{'date': item['date'], 'value': item['value']} for item in historical_data.get('3day', [])])},
    weekly: {json.dumps([{'date': item['date'], 'value': item['value']} for item in historical_data.get('weekly', [])])}
}};

// 初始化汇总趋势图
function initSummaryChart() {{
    var chartDom = document.getElementById('summary-chart');
    if (!chartDom) return;
    
    var myChart = echarts.init(chartDom);
    
    // 准备数据
    var dates = [];
    var dailyData = [];
    var threeDayData = [];
    var weeklyData = [];
    
    // 收集所有日期
    var allDates = new Set();
    for (var period in summaryChartData) {{
        summaryChartData[period].forEach(function(item) {{
            allDates.add(item.date);
        }});
    }}
    
    // 排序日期
    dates = Array.from(allDates).sort();
    
    // 创建数据映射
    var dataMap = {{}};
    for (var period in summaryChartData) {{
        dataMap[period] = {{}};
        summaryChartData[period].forEach(function(item) {{
            dataMap[period][item.date] = item.value;
        }});
    }}
    
    // 填充数据
    dates.forEach(function(date) {{
        dailyData.push(dataMap.daily && dataMap.daily[date] !== undefined ? dataMap.daily[date] : null);
        threeDayData.push(dataMap['3day'] && dataMap['3day'][date] !== undefined ? dataMap['3day'][date] : null);
        weeklyData.push(dataMap.weekly && dataMap.weekly[date] !== undefined ? dataMap.weekly[date] : null);
    }});
    
    var option = {{
        title: {{
            text: '三周期满足数趋势图',
            left: 'center',
            textStyle: {{
                fontSize: 16,
                fontWeight: 'normal'
            }}
        }},
        tooltip: {{
            trigger: 'axis',
            formatter: function(params) {{
                var result = params[0].axisValue + '<br/>';
                params.forEach(function(param) {{
                    var value = param.value;
                    if (value !== null && value !== undefined) {{
                        var periodName = param.seriesName;
                        var color = param.color;
                        result += '<span style=\"display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;background-color:' + color + '\"></span>';
                        result += periodName + ': ' + parseFloat(value).toFixed(2) + '<br/>';
                    }}
                }});
                return result;
            }}
        }},
        legend: {{
            data: ['日线满足数', '三日满足数', '周线满足数/4'],
            top: 30,
            textStyle: {{
                fontSize: 12
            }}
        }},
        grid: {{
            left: '3%',
            right: '4%',
            bottom: '3%',
            top: 70,
            containLabel: true
        }},
        xAxis: {{
            type: 'category',
            boundaryGap: false,
            data: dates,
            axisLabel: {{
                fontSize: 10,
                rotate: 45
            }}
        }},
        yAxis: {{
            type: 'value',
            max: 250,
            axisLabel: {{
                fontSize: 11
            }},
            splitLine: {{
                lineStyle: {{
                    type: 'dashed'
                }}
            }}
        }},
        series: [
            {{
                name: '日线满足数',
                type: 'line',
                smooth: true,
                lineStyle: {{
                    width: 2,
                    color: '#ff0000'  // 红色
                }},
                data: dailyData
            }},
            {{
                name: '三日满足数',
                type: 'line',
                smooth: true,
                lineStyle: {{
                    width: 2,
                    color: '#000080'  // 深蓝色
                }},
                data: threeDayData
            }},
            {{
                name: '周线满足数/4',
                type: 'line',
                smooth: true,
                lineStyle: {{
                    width: 2,
                    color: '#ffff00'  // 黄色
                }},
                data: weeklyData
            }},
            // 固定横线：15（紫色虚线）
            {{
                name: '15线',
                type: 'line',
                markLine: {{
                    silent: true,
                    lineStyle: {{
                        type: 'dashed',
                        color: '#800080'  // 紫色
                    }},
                    data: [{{yAxis: 15}}]
                }}
            }},
            // 固定横线：50（白色虚线）
            {{
                name: '50线',
                type: 'line',
                markLine: {{
                    silent: true,
                lineStyle: {{
                    type: 'dashed',
                    color: '#000000'  // 黑色
                }},
                    data: [{{yAxis: 50}}]
                }}
            }},
            // 固定横线：100（黄色虚线）
            {{
                name: '100线',
                type: 'line',
                markLine: {{
                    silent: true,
                    lineStyle: {{
                        type: 'dashed',
                        color: '#ffff00'  // 黄色
                    }},
                    data: [{{yAxis: 100}}]
                }}
            }}
        ]
    }};
    
    myChart.setOption(option);
    
    // 响应窗口大小变化
    window.addEventListener('resize', function() {{
        myChart.resize();
    }});
}}

var currentTabPeriod = 'daily';
function switchTab(name, event) {{
  // 获取被点击的tab元素
  var clickedTab = event ? event.currentTarget : this;
  
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
  
  // 确保目标tab内容元素存在再添加active类
  var targetTab = document.getElementById('tab-' + name);
  if (targetTab) {{
    targetTab.classList.add('active');
  }}
  
  // 为被点击的tab元素添加active类
  if (clickedTab) {{
    clickedTab.classList.add('active');
  }}
  currentTabPeriod = name;
  // 切换 Tab 时初始化该 Tab 的图表和表格
  if (typeof window['initChartForTab_' + name] === 'function') {{
    window['initChartForTab_' + name]();
  }}
}}

function highlightSector(sectorCode) {{
  // 移除之前的高亮
  document.querySelectorAll('.sector-highlighted').forEach(el => {{
    el.classList.remove('sector-highlighted');
  }});

  // 高亮所有匹配的板块
  const cells = document.querySelectorAll('[data-sector-code="' + sectorCode + '"]');
  cells.forEach(cell => {{
    cell.classList.add('sector-highlighted');
  }});
}}

// 页面加载完成后初始化汇总图表和各Tab图表
window.addEventListener('DOMContentLoaded', function() {{
    // 初始化汇总趋势图
    if (typeof initSummaryChart === 'function') {{
        initSummaryChart();
    }}
    
    // 初始化日线趋势图
    if (typeof initChartForTab_daily === 'function') {{
        initChartForTab_daily();
    }}
}});
</script>
</body>
</html>"""

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"[HTML] 报告已生成: {output_file}")


def main(date=None, full_mode=False, export_excel=False, use_intraday=False, run_id=None):
    """
    date: 指定报告日期，YYYYMMDD 或 YYYY-MM-DD 格式
          三 Tab 各自显示 daily/3day/weekly 三个周期
    """
    print("=" * 70)
    print("HTML 报告生成器 - V2.0（三周期版）")
    print("=" * 70)

    # 检查数据库文件
    if not check_database_file():
        return

    # 检查并创建报告目录
    if not check_report_directory():
        return

    # 设置输出文件路径
    global OUTPUT_FILE
    report_date_str = date.replace('-', '') if date else datetime.now().strftime('%Y%m%d')
    OUTPUT_FILE = f"{REPORT_DIR}/concept_report_{report_date_str}.html"
    print(f"数据库文件: {DB_FILE}")
    print(f"报告目录: {REPORT_DIR}")
    print(f"输出文件: {OUTPUT_FILE}")

    conn = sqlite3.connect(DB_FILE)

    # 解析日期 + 查找三个周期的 run_id
    if date:
        date_normalized = date.replace('-', '')
        run_id_daily  = get_run_id_for_date(conn, date_normalized, 'daily')
        run_id_3day   = get_run_id_for_date(conn, date_normalized, '3day')
        run_id_weekly = get_run_id_for_date(conn, date_normalized, 'weekly')
        periods_found = [p for p, r in [('日线', run_id_daily), ('三日', run_id_3day), ('周线', run_id_weekly)] if r]
        print(f"报告日期: {date_normalized}，可用周期: {', '.join(periods_found) if periods_found else '无数据'}")
        if not (run_id_daily or run_id_3day or run_id_weekly):
            print("[ERROR] 该日期没有任何周期的数据，请先运行 backfill")
            conn.close()
            return
    elif run_id:
        # 向后兼容旧 --run-id 参数：从中提取日期
        date_normalized = run_id.split('_')[0] if '_' in run_id else run_id
        run_id_daily  = get_run_id_for_date(conn, date_normalized, 'daily')
        run_id_3day   = get_run_id_for_date(conn, date_normalized, '3day')
        run_id_weekly = get_run_id_for_date(conn, date_normalized, 'weekly')
        periods_found = [p for p, r in [('日线', run_id_daily), ('三日', run_id_3day), ('周线', run_id_weekly)] if r]
        print(f"报告日期: {date_normalized}（从 --run-id 兼容参数解析），可用周期: {', '.join(periods_found) if periods_found else '无数据'}")
        if not (run_id_daily or run_id_3day or run_id_weekly):
            print("[ERROR] 该日期没有任何周期的数据，请先运行 backfill")
            conn.close()
            return
    else:
        # 默认：取最新日期的三个周期
        latest = get_latest_run_id(conn)
        if latest is None:
            print("[ERROR] 数据库无运行记录，请先运行概念板块分析脚本")
            conn.close()
            return
        date_normalized = latest.split('_')[0]
        run_id_daily  = get_run_id_for_date(conn, date_normalized, 'daily')
        run_id_3day   = get_run_id_for_date(conn, date_normalized, '3day')
        run_id_weekly = get_run_id_for_date(conn, date_normalized, 'weekly')
        print(f"默认使用最新日期: {date_normalized}，可用周期: 日线={bool(run_id_daily)}, 三日={bool(run_id_3day)}, 周线={bool(run_id_weekly)}")

    print(f"数据模式: {'全部数据' if full_mode else '最近三个月'}")
    print(f"板块涨幅: {'盘中实时' if use_intraday else '收盘数据'}")

    # 生成 HTML 报告（三个 Tab 各用各的 run_id）
    generate_html(conn, run_id_daily, run_id_3day, run_id_weekly,
                  OUTPUT_FILE, REPORT_DIR, full_mode=full_mode, use_intraday=use_intraday)

    # 导出 Excel 文件（可选）
    if export_excel:
        print("\n开始导出 Excel 文件...")
        default_run = run_id_daily or run_id_3day or run_id_weekly
        all_sectors = load_sector_stats(conn, default_run)
        export_stocks_to_excel(conn, default_run, all_sectors, REPORT_DIR)

        # 加载趋势数据用于导出（使用周线数据）
        all_dates, _, _ = load_trend_data(conn, period='weekly', top_n_sectors=20)
        if all_dates:
            daily_top20 = load_daily_top20(conn, 'weekly', all_dates)
            export_trend_to_excel(conn, all_dates, daily_top20, REPORT_DIR, default_run)

    conn.close()

    print(f"\n[SUCCESS] HTML 报告已生成: {OUTPUT_FILE}")
    print(f"[INFO]    版本: V2.0（三周期版）")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Generate concept sector HTML report (V2.0 - three periods)')
    parser.add_argument('--date', type=str,
                        help='Report date: YYYYMMDD or YYYY-MM-DD (default: latest date in DB)')
    parser.add_argument('--run-id', type=str,
                        help='[Deprecated] Run ID like 20260402_HIST_WEEKLY. Use --date instead.')
    parser.add_argument('--full', action='store_true',
                        help='Load all historical data (default: last 3 months only)')
    parser.add_argument('--export', action='store_true',
                        help='Export Excel files')
    parser.add_argument('--intraday', action='store_true',
                        help='Use intraday real-time data (from TDX cache)')
    args = parser.parse_args()
    main(date=args.date, full_mode=args.full, export_excel=args.export,
         use_intraday=args.intraday, run_id=args.run_id)