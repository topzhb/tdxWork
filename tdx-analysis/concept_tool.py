#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
concept_tool.py — 妖线概念板块分析工具
版本：V2.0  更新日期：2026-04-02

V2.0 更新：
- backfill 同时计算日线/三日/周线三个周期
- 数据库新增 period 字段（'daily'|'3day'|'weekly'）
- run_id 格式：{YYYYMMDD_HIST}_{PERIOD}

用法：
  python concept_tool.py backfill [选项]
  python concept_tool.py export   [选项]
  python concept_tool.py info

子命令：
  backfill  历史数据回溯（按日期遍历，每个工作日同时写入 daily/3day/weekly 三个周期）
  export    从数据库导出 Excel
  info      查看数据库当前数据概况

示例：
  # 回溯最近30天（逐日遍历，每个工作日补算缺失的周期）
  python concept_tool.py backfill --days 30

  # 回溯指定日期范围
  python concept_tool.py backfill --start 2026-01-01 --end 2026-03-20

  # 导出预设板块（tool_config.py 中的 SECTOR_PRESETS）到默认目录
  python concept_tool.py export

  # 导出指定板块
  python concept_tool.py export --sectors CPO概念 固态电池 芯片

  # 导出所有板块
  python concept_tool.py export --all

  # 导出指定日期范围
  python concept_tool.py export --start 2025-10-01 --end 2026-03-20

  # 导出并指定输出文件
  python concept_tool.py export --output D:/reports/my_report.xlsx

  # 查看数据库概况
  python concept_tool.py info
"""

import os
import sys
import struct
import math
import datetime
import sqlite3
import argparse
from collections import defaultdict

# ============================================================
# 加载配置
# ============================================================
try:
    from tool_config import (
        DB_FILE, CONCEPT_FILE, OUTPUT_DIR, VIPDOC_DIR,
        MIN_BARS, MIN_DAY_BARS, ANGLE_THRESH,
        EMA_PERIOD_X1, EMA_PERIOD_ZS1, ZS1_MULT,
        SECTOR_PRESETS
    )
except ImportError:
    print("[ERROR] 找不到 tool_config.py，请确保与 concept_tool.py 在同一目录")
    sys.exit(1)


# ============================================================
# 工具函数
# ============================================================
def log(msg, level="INFO"):
    prefix = {"INFO": "[INFO]", "OK": "[ OK ]", "WARN": "[WARN]", "ERR": "[ERR ]"}.get(level, "[INFO]")
    print(f"{prefix} {msg}")


def parse_date(s):
    try:
        return datetime.datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise argparse.ArgumentTypeError(f"日期格式错误，应为 YYYY-MM-DD，输入为: {s}")


# ============================================================
# 数据库初始化
# ============================================================
def init_db(conn):
    conn.executescript("""
    PRAGMA journal_mode=WAL;
    PRAGMA synchronous=NORMAL;

    CREATE TABLE IF NOT EXISTS t_sector (
        sector_code  TEXT PRIMARY KEY,
        sector_name  TEXT NOT NULL,
        created_at   TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS t_sector_stock (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        sector_code  TEXT NOT NULL,
        stock_code   TEXT NOT NULL,
        stock_name   TEXT,
        UNIQUE(sector_code, stock_code)
    );

    CREATE TABLE IF NOT EXISTS t_stock (
        stock_code   TEXT PRIMARY KEY,
        stock_name   TEXT,
        market       TEXT,
        updated_at   TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS t_stock_calc (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        run_id       TEXT NOT NULL,
        stock_code   TEXT NOT NULL,
        angle        REAL,
        zhusheng1    REAL,
        close_price  REAL,
        x_val        INTEGER,
        x1_ema9      REAL,
        is_satisfied INTEGER NOT NULL DEFAULT 0,
        calc_date    TEXT NOT NULL,
        UNIQUE(run_id, stock_code)
    );

    CREATE TABLE IF NOT EXISTS t_sector_stat (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        run_id          TEXT NOT NULL,
        sector_code     TEXT NOT NULL,
        rank_no         INTEGER,
        total_stocks    INTEGER,
        analyzed_count  INTEGER,
        satisfied_count INTEGER,
        satisfied_rate  REAL,
        calc_date       TEXT NOT NULL,
        UNIQUE(run_id, sector_code)
    );

    CREATE TABLE IF NOT EXISTS t_run_log (
        run_id          TEXT PRIMARY KEY,
        version         TEXT NOT NULL,
        total_sectors   INTEGER,
        total_stocks    INTEGER,
        unique_stocks   INTEGER,
        analyzed_count  INTEGER,
        satisfied_count INTEGER,
        satisfied_rate  REAL,
        run_time        TEXT NOT NULL,
        note            TEXT
    );

    CREATE TABLE IF NOT EXISTS t_daily_report (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        report_date     TEXT NOT NULL,
        run_id          TEXT NOT NULL,
        sector_code     TEXT NOT NULL,
        sector_name     TEXT NOT NULL,
        rank_no         INTEGER,
        total_stocks    INTEGER,
        analyzed_count  INTEGER,
        satisfied_count INTEGER,
        satisfied_rate  REAL,
        UNIQUE(report_date, sector_code)
    );

    CREATE INDEX IF NOT EXISTS idx_stock_calc_run    ON t_stock_calc(run_id);
    CREATE INDEX IF NOT EXISTS idx_stock_calc_stock  ON t_stock_calc(stock_code);
    CREATE INDEX IF NOT EXISTS idx_stock_calc_sat    ON t_stock_calc(run_id, is_satisfied);
    CREATE INDEX IF NOT EXISTS idx_sector_stat_run   ON t_sector_stat(run_id);
    CREATE INDEX IF NOT EXISTS idx_daily_report_date ON t_daily_report(report_date);
    CREATE INDEX IF NOT EXISTS idx_daily_report_code ON t_daily_report(sector_code);
    """)
    conn.commit()


# ============================================================
# 从数据库读取概念板块数据
# ============================================================
def read_sector_data_from_db(conn):
    """从数据库读取板块和股票数据"""
    c = conn.cursor()
    
    # 读取板块数据
    c.execute("SELECT sector_code, sector_name FROM t_sector ORDER BY sector_code")
    sector_rows = c.fetchall()
    
    if not sector_rows:
        raise ValueError("数据库中没有板块数据，请先运行 update_concept_data_with_filter.py 导入概念板块数据")
    
    sector_names = {}
    for s_code, s_name in sector_rows:
        sector_names[s_code] = s_name
    
    # 读取板块-股票关系
    sectors = defaultdict(list)
    stock_names = {}
    unique_stocks = {}
    
    c.execute("""
        SELECT ss.sector_code, ss.stock_code, s.stock_name, s.market
        FROM t_sector_stock ss
        JOIN t_stock s ON ss.stock_code = s.stock_code
        ORDER BY ss.sector_code, ss.stock_code
    """)
    
    for s_code, st_code, st_name, market in c.fetchall():
        sectors[s_code].append(st_code)
        stock_names[st_code] = st_name
        unique_stocks[st_code] = (st_name, market)
    
    return sector_names, sectors, unique_stocks

# 原import_concept_sectors函数已注释掉，使用独立的update_concept_data_with_filter.py更新概念数据
# def import_concept_sectors(conn, concept_file):
#     # ... 原函数代码已注释掉 ...


# ============================================================
# 计算函数
# ============================================================
def ema(values, period):
    if not values:
        return []
    result = [values[0]]
    mult = 2.0 / (period + 1)
    for v in values[1:]:
        result.append(v * mult + result[-1] * (1 - mult))
    return result


def read_tdx_day_file(filepath, target_date=None):
    if not os.path.exists(filepath):
        return None
    try:
        records = []
        with open(filepath, 'rb') as f:
            while True:
                rec = f.read(32)
                if len(rec) < 32:
                    break
                date_int, o, h, l, cl, vol, amt, _ = struct.unpack('<7I4s', rec)
                try:
                    y = date_int // 10000
                    m = (date_int % 10000) // 100
                    d = date_int % 100
                    rec_date = datetime.date(y, m, d)
                except Exception:
                    continue
                if target_date and rec_date > target_date:
                    continue
                records.append([date_int, o/100.0, h/100.0, l/100.0, cl/100.0, vol, amt])
        return records if records else None
    except Exception:
        return None


def get_week_key(date_int):
    try:
        y = date_int // 10000
        m = (date_int % 10000) // 100
        d = date_int % 100
        iso = datetime.date(y, m, d).isocalendar()
        return (iso[0], iso[1])
    except Exception:
        return None


def convert_to_weekly_natural(day_data):
    if not day_data or len(day_data) < 5:
        return None
    weeks = {}
    for bar in day_data:
        key = get_week_key(bar[0])
        if key is None:
            continue
        weeks.setdefault(key, []).append(bar)
    weekly = []
    for key in sorted(weeks):
        wb = sorted(weeks[key], key=lambda x: x[0])
        weekly.append([
            wb[-1][0], wb[0][1],
            max(b[2] for b in wb),
            min(b[3] for b in wb),
            wb[-1][4],
            sum(b[5] for b in wb),
            sum(b[6] for b in wb),
        ])
    return weekly


def convert_to_daily_raw(day_data):
    """日线：不聚合，直接用原始日线数据"""
    if not day_data or len(day_data) < 5:
        return None
    # 返回格式同周线（直接用日线数据）
    return [list(bar) for bar in day_data]


def convert_to_3day_natural(day_data):
    """三日线：每3个连续交易日合成一根K线，从头开始不跳步"""
    if not day_data or len(day_data) < 5:
        return None
    result = []
    for i in range(0, len(day_data), 3):
        group = day_data[i:i+3]
        if not group:
            continue
        result.append([
            group[-1][0],           # date: 最后一天
            group[0][1],            # open: 第一天开盘
            max(b[2] for b in group),    # high: 3天最高
            min(b[3] for b in group),    # low:  3天最低
            group[-1][4],           # close: 最后一天收盘
            sum(b[5] for b in group),   # vol
            sum(b[6] for b in group),   # amount
        ])
    return result


def calculate_yao_xian(weekly_data, stock_code):
    if not weekly_data or len(weekly_data) < MIN_BARS:
        return None, None, None, None, None
    try:
        close = [d[4] for d in weekly_data]
        H21 = 1 if stock_code.startswith('688') else 0
        H20 = 1 if stock_code.startswith('30')  else 0
        H11 = 1 if stock_code.startswith('60')  else 0
        H10 = 1 if stock_code.startswith('00')  else 0
        X   = (H11 + H10 > 0) * 1 + (H21 + H20 > 0) * 2

        x1  = ema(close, EMA_PERIOD_X1)
        zs1 = ema([v * ZS1_MULT for v in x1], EMA_PERIOD_ZS1)

        if len(zs1) >= 2 and zs1[-2] > 0:
            ratio = (zs1[-1] / zs1[-2] - 1) * 100
            angle = math.degrees(math.atan(ratio))
        else:
            angle = 0.0

        return angle, zs1[-1], close[-1], X, x1[-1]
    except Exception:
        return None, None, None, None, None


def is_st_stock(stock_name):
    """判断是否为ST股票"""
    if stock_name is None:
        return False
    # 检查是否包含ST (包括ST、*ST、ST*等情况)
    stock_name_upper = stock_name.upper()
    return 'ST' in stock_name_upper


def check_condition(angle, close, zhusheng, x):
    try:
        return (angle is not None and angle >= ANGLE_THRESH and
                close is not None and close >= zhusheng and
                x     is not None and x     >  0)
    except Exception:
        return False


def check_3day_consecutive_satisfied(day_data, stock_code):
    """检查连续3个交易日是否都满足妖线条件
    
    Args:
        day_data: 日线数据列表，格式为 [[date_int, open, high, low, close, vol, amt], ...]
        stock_code: 股票代码
    
    Returns:
        bool: True表示连续3天都满足，False表示不满足
    """
    if not day_data or len(day_data) < MIN_DAY_BARS:
        return False
    
    # 我们需要最近3个交易日的数据
    # 确保数据是按时间顺序的（最早的在前，最新的在后）
    if len(day_data) < 3:
        return False
    
    # 我们需要计算每一天的妖线条件
    # 但妖线计算需要一定数量的历史数据
    # 对于每一天，我们都用截止到那一天的日线数据来计算
    last_3_satisfied = []
    
    # 对于最近3天的每一天
    for i in range(-3, 0):
        # 获取截止到第i天的日线数据（包含第i天）
        data_up_to_day = day_data[:len(day_data) + i + 1]
        
        # 计算妖线参数
        angle, zs1, close, x_val, x1_ema9 = calculate_yao_xian(data_up_to_day, stock_code)
        if angle is None:
            return False  # 任何一天计算失败都返回False
        
        # 判断这一天是否满足条件
        satisfied = check_condition(angle, close, zs1, x_val)
        last_3_satisfied.append(satisfied)
    
    # 必须连续3天都满足
    return all(last_3_satisfied)


# ============================================================
# 回溯：单日分析
# ============================================================
def analyze_for_date(conn, target_date, sector_names, sectors, unique_stocks,
                     period='weekly', verbose=True):
    """
    单日单周期分析。
    period: 'daily' | '3day' | 'weekly'
    """
    c         = conn.cursor()
    calc_date = target_date.strftime("%Y-%m-%d")
    run_id    = f"{target_date.strftime('%Y%m%d_HIST')}_{period.upper()}"
    total_uniq = len(unique_stocks)

    # 周期 → K线聚合函数
    agg_map = {
        'daily':  convert_to_daily_raw,
        '3day':   convert_to_daily_raw,  # 三日周期也使用日线数据，但判断逻辑不同
        'weekly': convert_to_weekly_natural,
    }
    agg_func = agg_map.get(period, convert_to_weekly_natural)
    period_label = {'daily': '日线', '3day': '三日', 'weekly': '周线'}.get(period, period)

    if verbose:
        log(f"回溯 {calc_date} [{period_label}] ...")

    # Step-1: 逐股计算
    calc_rows   = []
    calc_result = {}
    done        = 0

    for stock_code, (stock_name, market) in unique_stocks.items():
        if market == 'other':
            continue

        # 过滤ST股票
        if is_st_stock(stock_name):
            if verbose and done < 10:  # 只打印前10个被过滤的ST股票,避免输出过多
                log(f"  跳过ST股票: {stock_name} ({stock_code})")
            done += 1
            continue

        day_file = os.path.join(VIPDOC_DIR, market, 'lday', f'{market}{stock_code}.day')
        day_data = read_tdx_day_file(day_file, target_date)
        if not day_data or len(day_data) < MIN_DAY_BARS:
            done += 1
            continue

        agg_data = agg_func(day_data)
        if not agg_data or len(agg_data) < 30:
            done += 1
            continue

        angle, zs1, close, x_val, x1_ema9 = calculate_yao_xian(agg_data, stock_code)
        if angle is None:
            done += 1
            continue

        # 计算当日涨幅（使用日线数据最后两条）
        daily_change = 0.0
        if len(day_data) >= 2:
            prev_close = day_data[-2][4]  # 昨日收盘价
            curr_close = day_data[-1][4]  # 今日收盘价
            if prev_close > 0:
                daily_change = (curr_close - prev_close) / prev_close * 100

        # 根据周期使用不同的判断逻辑
        if period == '3day':
            # 三日周期：检查连续3天是否都满足妖线条件
            is_ok = 1 if check_3day_consecutive_satisfied(day_data, stock_code) else 0
        else:
            # 日线和周线：使用常规判断
            is_ok = 1 if check_condition(angle, close, zs1, x_val) else 0
        calc_result[stock_code] = {'angle': angle, 'zhusheng1': zs1,
                                   'close': close, 'x_val': x_val, 'is_satisfied': is_ok,
                                   'daily_change': daily_change}
        calc_rows.append((
            run_id, stock_code,
            round(angle,    6),
            round(zs1,      6) if zs1     is not None else None,
            round(close,    4) if close   is not None else None,
            x_val,
            round(x1_ema9,  6) if x1_ema9 is not None else None,
            is_ok, calc_date,
            round(daily_change, 4),
            period
        ))
        done += 1
        if verbose and done % 500 == 0:
            print(f"      已计算 {done}/{total_uniq} 只...")

    c.executemany("""
        INSERT OR REPLACE INTO t_stock_calc
            (run_id, stock_code, angle, zhusheng1, close_price,
             x_val, x1_ema9, is_satisfied, calc_date, daily_change, period)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, calc_rows)
    conn.commit()
    if verbose:
        log(f"  t_stock_calc 写入 {len(calc_rows)} 条")

    # Step-2: 板块聚合
    results         = []
    total_analyzed  = 0
    total_satisfied = 0
    total_all_stocks = sum(len(v) for v in sectors.values())

    for sector_code, stocks in sectors.items():
        analyzed_cnt  = 0
        satisfied_cnt = 0
        for st_code in stocks:
            res = calc_result.get(st_code)
            if res is None:
                continue
            analyzed_cnt  += 1
            satisfied_cnt += res['is_satisfied']

        sat_rate = (satisfied_cnt / analyzed_cnt * 100) if analyzed_cnt else 0.0
        results.append({
            'sector_code':    sector_code,
            'sector_name':    sector_names.get(sector_code, '未知'),
            'total_stocks':   len(stocks),
            'analyzed_count': analyzed_cnt,
            'satisfied_count': satisfied_cnt,
            'satisfied_rate': sat_rate,
        })
        total_analyzed  += analyzed_cnt
        total_satisfied += satisfied_cnt

    results.sort(key=lambda r: r['satisfied_count'], reverse=True)
    for rank, r in enumerate(results, 1):
        r['rank'] = rank

    c.executemany("""
        INSERT OR REPLACE INTO t_sector_stat
            (run_id, sector_code, rank_no, total_stocks,
             analyzed_count, satisfied_count, satisfied_rate, calc_date, period)
        VALUES (?,?,?,?,?,?,?,?,?)
    """, [(run_id, r['sector_code'], r['rank'], r['total_stocks'],
           r['analyzed_count'], r['satisfied_count'],
           round(r['satisfied_rate'], 4), calc_date, period) for r in results])
    conn.commit()
    if verbose:
        log(f"  t_sector_stat 写入 {len(results)} 条")

    # Step-3: 每日报告快照
    c.executemany("""
        INSERT OR REPLACE INTO t_daily_report
            (report_date, run_id, sector_code, sector_name,
             rank_no, total_stocks, analyzed_count,
             satisfied_count, satisfied_rate, period)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, [(calc_date, run_id, r['sector_code'], r['sector_name'],
           r['rank'], r['total_stocks'], r['analyzed_count'],
           r['satisfied_count'], round(r['satisfied_rate'], 4), period) for r in results])
    conn.commit()
    if verbose:
        log(f"  t_daily_report 写入 {len(results)} 条")

    # Step-4: 运行记录
    dedup_satisfied = c.execute(
        "SELECT COUNT(*) FROM t_stock_calc WHERE run_id=? AND is_satisfied=1 AND period=?",
        (run_id, period)
    ).fetchone()[0]
    dedup_analyzed  = len(calc_rows)
    overall_rate    = dedup_satisfied / dedup_analyzed * 100 if dedup_analyzed else 0.0

    c.execute("""
        INSERT OR REPLACE INTO t_run_log
            (run_id, version, total_sectors, total_stocks, unique_stocks,
             analyzed_count, satisfied_count, satisfied_rate, run_time, note)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (run_id, 'V2.0-回溯', len(results), total_all_stocks, len(calc_rows),
          dedup_analyzed, dedup_satisfied, round(overall_rate, 4),
          datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
          f'历史回溯：{calc_date} [{period_label}]'))
    conn.commit()

    if verbose:
        log(f"  {calc_date}[{period_label}] 完成：分析 {dedup_analyzed} 只，满足 {dedup_satisfied} 只，满足率 {overall_rate:.2f}%", "OK")

    return dedup_satisfied, overall_rate


# ============================================================
# 子命令：backfill
# ============================================================
def cmd_backfill(args):
    print("=" * 70)
    print("历史数据回溯")
    print("=" * 70)

    # 确定日期范围
    end_date   = args.end   or datetime.date.today()
    if args.start:
        start_date = args.start
    elif args.days:
        start_date = end_date - datetime.timedelta(days=args.days - 1)
    else:
        print("[ERROR] 请指定 --start 或 --days")
        sys.exit(1)

    # 已回溯日期（跳过重复）—— 按 (report_date, period) 组合判断
    conn = sqlite3.connect(DB_FILE)
    init_db(conn)
    existing = set(
        (row[0], row[1]) for row in conn.execute(
            "SELECT DISTINCT report_date, period FROM t_daily_report"
        ).fetchall()
    )

    if args.skip_existing:
        log(f"数据库中已有 {len(existing)} 个(日期,周期)记录，将跳过已有组合")
    else:
        log(f"注意：--no-skip 模式，已有记录将被覆盖")

    log(f"回溯范围：{start_date} ~ {end_date}")

    PERIODS = ['daily', '3day', 'weekly']

    # 统计待回溯天数（以"日期有任意一个周期缺失"为条件）
    pending_dates = []
    cur = start_date
    while cur <= end_date:
        if cur.weekday() < 5:  # 跳过周末
            date_str = cur.strftime("%Y-%m-%d")
            if args.skip_existing and all((date_str, p) in existing for p in PERIODS):
                pass  # 三个周期都已存在才跳过
            else:
                pending_dates.append(cur)
        cur += datetime.timedelta(days=1)

    log(f"需要回溯 {len(pending_dates)} 个交易日")
    if not pending_dates:
        log("没有需要回溯的日期，退出", "OK")
        conn.close()
        return

    # 从数据库读取板块数据
    log("从数据库读取板块数据...")
    try:
        sector_names, sectors, unique_stocks = read_sector_data_from_db(conn)
        log(f"板块数量: {len(sector_names)}，股票数量: {len(unique_stocks)}", "OK")
    except ValueError as e:
        log(f"读取数据库失败: {e}", "ERR")
        log("请先运行 update_concept_data_with_filter.py 导入概念板块数据", "INFO")
        conn.close()
        return

    # 逐日逐周期回溯（daily / 3day / weekly 同时计算）
    period_label_map = {'daily': '日线', '3day': '三日', 'weekly': '周线'}
    print()
    import sys
    success, failed = 0, 0
    for i, target_date in enumerate(pending_dates, 1):
        date_str = target_date.strftime("%Y-%m-%d")
        print(f"[{i}/{len(pending_dates)}] {target_date} ", end="", flush=True)
        for period in PERIODS:
            # 跳过已存在的周期
            if args.skip_existing and (date_str, period) in existing:
                print(f"[-]", end="", flush=True)
                continue
            try:
                pbar = period_label_map[period]
                print(f"[{pbar}]", end="", flush=True)
                analyze_for_date(conn, target_date, sector_names, sectors, unique_stocks,
                                 period=period, verbose=False)
            except KeyboardInterrupt:
                log("用户中断", "WARN")
                conn.close()
                return
            except Exception as e:
                log(f"回溯 {target_date}[{period}] 失败: {e}", "ERR")
                failed += 1
        print(" OK", flush=True)
        success += 3   # 每天3个周期都成功才算成功1天

    conn.close()
    print()
    print("=" * 70)
    log(f"回溯完成：{len(pending_dates)} 天 × 3 周期，成功", "OK")
    print("=" * 70)


# ============================================================
# 子命令：export
# ============================================================
def _adjust_col_width(worksheet):
    for column in worksheet.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 50)


def cmd_export(args):
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log("缺少依赖：pip install pandas openpyxl", "ERR")
        sys.exit(1)

    print("=" * 70)
    print("板块数据导出")
    print("=" * 70)

    conn = sqlite3.connect(DB_FILE)

    # 确定板块列表
    if args.all:
        rows = conn.execute(
            "SELECT DISTINCT sector_code, sector_name FROM t_daily_report ORDER BY sector_name"
        ).fetchall()
        target_sectors = {name: code for code, name in rows}
        log(f"导出全部板块，共 {len(target_sectors)} 个")
    elif args.sectors:
        # 用户指定板块名，从预设中查找代码；未找到则尝试从数据库查
        target_sectors = {}
        for name in args.sectors:
            if name in SECTOR_PRESETS:
                target_sectors[name] = SECTOR_PRESETS[name]
            else:
                row = conn.execute(
                    "SELECT sector_code FROM t_daily_report WHERE sector_name=? LIMIT 1", (name,)
                ).fetchone()
                if row:
                    target_sectors[name] = row[0]
                else:
                    log(f"未找到板块 '{name}'，跳过", "WARN")
        log(f"指定板块：{list(target_sectors.keys())}")
    else:
        target_sectors = SECTOR_PRESETS
        log(f"使用预设板块：{list(target_sectors.keys())}")

    if not target_sectors:
        log("没有可导出的板块", "ERR")
        conn.close()
        sys.exit(1)

    # 确定日期范围
    all_dates_rows = conn.execute(
        "SELECT DISTINCT report_date FROM t_daily_report ORDER BY report_date ASC"
    ).fetchall()
    if not all_dates_rows:
        log("数据库中没有数据", "ERR")
        conn.close()
        sys.exit(1)

    all_dates = [r[0] for r in all_dates_rows]
    if args.start:
        all_dates = [d for d in all_dates if d >= args.start.strftime("%Y-%m-%d")]
    if args.end:
        all_dates = [d for d in all_dates if d <= args.end.strftime("%Y-%m-%d")]

    # 过滤掉周末（数据库中理论上已经是交易日，但以防万全）
    def is_trading_day(s):
        try:
            return datetime.datetime.strptime(s, "%Y-%m-%d").weekday() < 5
        except Exception:
            return True

    trading_dates = [d for d in all_dates if is_trading_day(d)]
    log(f"数据范围：{trading_dates[0]} ~ {trading_dates[-1]}，共 {len(trading_dates)} 个交易日")

    # 导出各板块数据
    export_data = {}

    for sector_name, sector_code in target_sectors.items():
        log(f"导出板块：{sector_name} ({sector_code})")
        rows = conn.execute("""
            SELECT
                ds.report_date,
                ds.sector_code,
                ds.sector_name,
                ds.satisfied_count,
                ds.rank_no,
                ds.satisfied_rate,
                (SELECT satisfied_count FROM t_run_log
                 WHERE run_id = (
                   SELECT run_id FROM t_daily_report
                   WHERE report_date = ds.report_date LIMIT 1
                 )
                ) as total_satisfied
            FROM t_daily_report ds
            WHERE ds.sector_code = ?
              AND ds.report_date >= ?
              AND ds.report_date <= ?
            ORDER BY ds.report_date ASC
        """, (sector_code,
              trading_dates[0],
              trading_dates[-1])).fetchall()

        if not rows:
            log(f"  未找到数据，跳过", "WARN")
            continue

        records = []
        for r in rows:
            report_date, s_code, s_name, sat_cnt, rank, rate, total_sat = r
            total_rate = (sat_cnt / total_sat * 100) if total_sat and total_sat > 0 else 0
            records.append({
                '日期':          report_date,
                '板块代码':      s_code,
                '板块名称':      s_name,
                '满足数':        sat_cnt,
                '排名':          rank,
                '板块满足率(%)': round(rate, 2) if rate else 0,
                '总满足数':      total_sat or 0,
                '总比满足率(%)': round(total_rate, 2),
            })
        export_data[sector_name] = records
        log(f"  导出 {len(records)} 条", "OK")

    # 总体趋势
    log("导出总体趋势...")
    global_rows = conn.execute("""
        SELECT
            dr.report_date,
            COUNT(DISTINCT dr.sector_code) as sector_count,
            rl.satisfied_count             as total_satisfied,
            rl.satisfied_rate              as overall_rate
        FROM t_daily_report dr
        LEFT JOIN t_run_log rl ON rl.run_id = (
            SELECT run_id FROM t_daily_report
            WHERE report_date = dr.report_date LIMIT 1
        )
        WHERE dr.report_date >= ? AND dr.report_date <= ?
        GROUP BY dr.report_date
        ORDER BY dr.report_date ASC
    """, (trading_dates[0], trading_dates[-1])).fetchall()

    global_records = []
    for r in global_rows:
        date_str, sector_cnt, total_sat, overall_rate = r
        if is_trading_day(date_str):
            global_records.append({
                '日期':             date_str,
                '参与板块数':        sector_cnt,
                '总满足数':          total_sat or 0,
                '整体满足率(%)':     round(overall_rate, 2) if overall_rate else 0,
            })
    export_data['总体趋势'] = global_records
    log(f"  导出 {len(global_records)} 条", "OK")

    # 导出信息表
    export_data['导出信息'] = [{
        '导出时间':    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '数据起始日期': trading_dates[0],
        '数据结束日期': trading_dates[-1],
        '交易日数':    len(trading_dates),
        '导出板块数':  len(target_sectors),
        '板块列表':   ', '.join(target_sectors.keys()),
    }]

    conn.close()

    # 确定输出路径
    if args.output:
        output_file = args.output
        os.makedirs(os.path.dirname(os.path.abspath(output_file)), exist_ok=True)
    else:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        timestamp   = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(OUTPUT_DIR, f"sectors_export_{timestamp}.xlsx")

    # 写入 Excel
    log(f"写入 Excel：{output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, data in export_data.items():
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            # 表头加粗
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", start_color="DDEEFF")
                cell.alignment = Alignment(horizontal="center")
            _adjust_col_width(ws)

    print()
    print("=" * 70)
    log(f"导出完成：{output_file}", "OK")
    log(f"工作表：{list(export_data.keys())}")
    print("=" * 70)


# ============================================================
# 子命令：info
# ============================================================
def cmd_info(args):
    if not os.path.exists(DB_FILE):
        log(f"数据库不存在：{DB_FILE}", "ERR")
        return

    conn = sqlite3.connect(DB_FILE)

    print("=" * 70)
    print("数据库概况")
    print("=" * 70)

    # 日期范围
    row = conn.execute(
        "SELECT MIN(report_date), MAX(report_date), COUNT(DISTINCT report_date) FROM t_daily_report"
    ).fetchone()
    if row and row[0]:
        print(f"  数据范围     : {row[0]} ~ {row[1]}")
        print(f"  总日期数     : {row[2]} 天")
    else:
        print("  t_daily_report 暂无数据")

    # 板块数
    n_sectors = conn.execute("SELECT COUNT(*) FROM t_sector").fetchone()[0]
    n_stocks  = conn.execute("SELECT COUNT(*) FROM t_stock").fetchone()[0]
    print(f"  板块总数     : {n_sectors}")
    print(f"  股票总数     : {n_stocks}")

    # 最近一次运行
    last_run = conn.execute(
        "SELECT run_id, run_time, satisfied_count, satisfied_rate FROM t_run_log ORDER BY run_time DESC LIMIT 1"
    ).fetchone()
    if last_run:
        print(f"\n  最近一次运行 : {last_run[1]}")
        print(f"  run_id       : {last_run[0]}")
        print(f"  满足数       : {last_run[2]}，满足率 {last_run[3]:.2f}%")

    # 预设板块最新数据
    print(f"\n  预设板块最新状态：")
    latest_date = conn.execute(
        "SELECT MAX(report_date) FROM t_daily_report"
    ).fetchone()[0]
    if latest_date:
        for name, code in SECTOR_PRESETS.items():
            row = conn.execute("""
                SELECT satisfied_count, rank_no, satisfied_rate
                FROM t_daily_report
                WHERE sector_code=? AND report_date=?
            """, (code, latest_date)).fetchone()
            if row:
                print(f"    {name:10s}  满足数={row[0]:3d}  排名={row[1]:4d}  板块满足率={row[2]:.2f}%")
            else:
                print(f"    {name:10s}  (无数据)")

    conn.close()
    print("=" * 70)


# ============================================================
# 命令行解析
# ============================================================
def build_parser():
    parser = argparse.ArgumentParser(
        prog="concept_tool",
        description="妖线概念板块分析工具 — 历史回溯 & Excel 导出",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""示例：
  python concept_tool.py backfill --days 30
  python concept_tool.py backfill --start 2026-01-01 --end 2026-03-20
  python concept_tool.py export
  python concept_tool.py export --sectors CPO概念 固态电池 芯片
  python concept_tool.py export --all --start 2025-10-01
  python concept_tool.py export --output D:/reports/my_report.xlsx
  python concept_tool.py info
"""
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    # ---------- backfill ----------
    bp = sub.add_parser("backfill", help="历史数据回溯（写入数据库）")
    bp.add_argument("--start", type=parse_date, metavar="YYYY-MM-DD",
                    help="回溯起始日期（与 --days 二选一）")
    bp.add_argument("--end",   type=parse_date, metavar="YYYY-MM-DD",
                    help="回溯结束日期（默认：今天）")
    bp.add_argument("--days",  type=int, metavar="N",
                    help="从结束日期向前 N 天（与 --start 二选一）")
    bp.add_argument("--skip-existing", dest="skip_existing",
                    action="store_true", default=True,
                    help="跳过数据库中已有的日期（默认开启）")
    bp.add_argument("--no-skip", dest="skip_existing",
                    action="store_false",
                    help="强制重新计算已有日期（覆盖）")
    bp.add_argument("--quiet", "-q", action="store_true",
                    help="减少输出（仅显示每日汇总）")

    # ---------- export ----------
    ep = sub.add_parser("export", help="从数据库导出 Excel")
    ep.add_argument("--sectors", nargs="+", metavar="板块名",
                    help="指定导出的板块名称（空格分隔，默认使用 tool_config.py 中的预设）")
    ep.add_argument("--all", action="store_true",
                    help="导出数据库中全部板块")
    ep.add_argument("--start", type=parse_date, metavar="YYYY-MM-DD",
                    help="数据起始日期")
    ep.add_argument("--end",   type=parse_date, metavar="YYYY-MM-DD",
                    help="数据结束日期")
    ep.add_argument("--output", "-o", metavar="文件路径",
                    help="输出 Excel 文件路径（默认：ConceptReport/sectors_export_时间戳.xlsx）")

    # ---------- info ----------
    sub.add_parser("info", help="查看数据库当前数据概况")

    return parser


def main():
    parser = build_parser()
    args   = parser.parse_args()

    if args.cmd == "backfill":
        cmd_backfill(args)
    elif args.cmd == "export":
        cmd_export(args)
    elif args.cmd == "info":
        cmd_info(args)


if __name__ == "__main__":
    main()
